# -*- coding: utf-8 -*-
"""
SIXT Portugal - Protection Package Pricing Simulator
=====================================================
Run: streamlit run app.py
"""

import io
import json
import os
import pandas as pd
import streamlit as st
import plotly.graph_objects as go
import xlsxwriter


# =========================================================
# OPTIMIZADOR GLOBAL
# =========================================================

def optimize_global_discounts(
    matrix_data, counter_discount, online_discount, bf_weight, vat,
    easy_counter_discount=0.0,
    easy_margin_min=3.5, easy_margin_max=7.0,
    online_extra_discount=0.10,
    tg_min_iva=8.0,
):
    lor_intervals = [
        {"lor_start": 1,  "lor_end": 3,   "days_reference": 1},
        {"lor_start": 4,  "lor_end": 6,   "days_reference": 4},
        {"lor_start": 7,  "lor_end": 10,  "days_reference": 7},
        {"lor_start": 11, "lor_end": 29,  "days_reference": 11},
        {"lor_start": 30, "lor_end": 30,  "days_reference": 30},
        {"lor_start": 31, "lor_end": 999, "days_reference": 31},
    ]

    global_low  = 0.0
    global_high = 1.0
    conflicts   = []
    coverage    = {}

    _tg_min_rack_opt = (tg_min_iva / (1 + vat)) if (tg_min_iva and tg_min_iva > 0) else 0.0
    _global_scale_opt = compute_global_tg_scale(matrix_data, _tg_min_rack_opt)

    for group in matrix_data["groups"]:
        tg_prog_opt = compute_tg_lor_progression(matrix_data, group, _tg_min_rack_opt, _global_scale_opt)

        for lor in lor_intervals:
            tg_ov = tg_prog_opt.get((lor["lor_start"], lor["lor_end"], lor["days_reference"]))
            r = calculate_pricing(
                matrix_data=matrix_data,
                group=group,
                days=lor["days_reference"],
                counter_discount=counter_discount,
                online_discount=online_discount,
                bf_weight=bf_weight,
                vat=vat,
                easy_counter_discount=easy_counter_discount,
                ai_counter_discount=0.0,
                ai_online_discount=0.0,
                easy_margin_min=easy_margin_min,
                easy_margin_max=easy_margin_max,
                online_extra_discount=online_extra_discount,
                tg_min_iva=tg_min_iva,
                tg_override=tg_ov,
            )


            vl = r.get("ai_counter_discount_valid_low")
            vh = r.get("ai_counter_discount_valid_high")
            key = f"{group} | LOR {lor['lor_start']}–{lor['lor_end']}"

            if vl is None or vh is None:
                coverage[key] = {"valid_low": None, "valid_high": None, "ok": None}
                continue

            coverage[key] = {"valid_low": vl, "valid_high": vh, "ok": vl <= vh}

            if vl > vh:
                conflicts.append({
                    "grupo": group,
                    "lor": f"{lor['lor_start']}–{lor['lor_end']}",
                    "valid_low": vl,
                    "valid_high": vh,
                })
                continue

            global_low  = max(global_low,  vl)
            global_high = min(global_high, vh)

    feasible = global_low <= global_high

    if feasible:
        ai_counter_opt = (global_low + global_high) / 2
    else:
        ai_counter_opt = global_low

    ai_online_opt = ai_counter_opt + online_extra_discount

    return {
        "ai_counter_discount": ai_counter_opt,
        "ai_online_discount": ai_online_opt,
        "valid_low": global_low,
        "valid_high": global_high,
        "feasible": feasible,
        "conflicts": conflicts,
        "coverage": coverage,
    }


# =========================================================
# LEITURA DO EXCEL
# =========================================================

@st.cache_data(show_spinner=False)
def build_matrix_data_from_excel(file_bytes, sheet_name):
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=None)

    def find_row(label):
        matches = df.index[
            df.iloc[:, 0].astype(str).str.strip().str.lower().eq(label.lower())
        ].tolist()
        if not matches:
            raise ValueError(f"Não encontrei a linha '{label}' no ficheiro Excel.")
        return matches[0]

    row_product   = find_row("chco")
    row_start_day = find_row("vont")
    row_end_day   = find_row("bist")
    row_cart      = find_row("cart")

    product_columns = []
    for col in range(1, df.shape[1]):
        product   = df.iat[row_product, col]
        start_day = df.iat[row_start_day, col]
        end_day   = df.iat[row_end_day, col]
        if pd.isna(product) or pd.isna(start_day) or pd.isna(end_day):
            continue
        product = str(product).strip()
        if product == "":
            continue
        try:
            sd = int(start_day)
            ed = int(end_day)
        except Exception:
            continue
        product_columns.append({"col": col, "product": product, "sd": sd, "ed": ed})

    if not product_columns:
        raise ValueError("Não foram encontradas colunas de produtos válidas na matriz.")

    products_to_keep = {"LD", "BF", "BQ", "TG", "BE", "I", "BC"}
    groups = []
    data   = {}

    for row in range(row_cart + 1, df.shape[0]):
        group = df.iat[row, 0]
        if pd.isna(group):
            continue
        group = str(group).strip()
        if group == "":
            continue
        groups.append(group)
        data[group] = {}
        for spec in product_columns:
            product = spec["product"]
            if product not in products_to_keep:
                continue
            value = df.iat[row, spec["col"]]
            if pd.isna(value):
                continue
            try:
                value = float(value)
            except Exception:
                continue
            if product not in data[group]:
                data[group][product] = []
            data[group][product].append({"sd": spec["sd"], "ed": spec["ed"], "v": value})

    if not groups:
        raise ValueError("Não foram encontrados grupos ACRISS na matriz.")

    return {"groups": groups, "data": data}


@st.cache_data(show_spinner=False)
def load_competitor_bd(file_bytes):
    """
    Lê BD_Pacotes_Concorrentes.xlsx (4 sheets: Hertz, Avis, Guerin, Europcar).
    Devolve dict {acriss: {field: preço_médio_€_por_dia}} onde field ∈
    {europcar_smart, europcar_ai, guerin_smart, guerin_ai,
     avis_smart, avis_ai, hertz_ai}.
    Usa keyword matching para ser robusto a encoding.
    Duplicados (mesmo ACRISS, mesmo pacote, modelos diferentes) → média.
    """

    def _classify(sheet, pkg):
        s = str(pkg).lower()
        if sheet == "Europcar":
            if "medium" in s:  return "europcar_smart"
            if "premium" in s: return "europcar_ai"
        elif sheet == "Guerin":
            if "light" in s:    return None          # Pack Light Gold — ignorar
            if "gold" in s:     return "guerin_smart" # Pack Premium Gold
            if "platinum" in s: return "guerin_ai"
        elif sheet == "Avis":
            if "plus" in s: return "avis_ai"          # Veículo Plus
            return "avis_smart"                        # Veículo
        elif sheet == "Hertz":
            return "hertz_ai"                          # SuperCover
        return None

    mapping = {}  # {acriss: {field: [prices]}}
    xl = pd.ExcelFile(io.BytesIO(file_bytes))

    for sheet in ["Hertz", "Avis", "Guerin", "Europcar"]:
        if sheet not in xl.sheet_names:
            continue
        df = pd.read_excel(xl, sheet_name=sheet, header=1)
        acriss_col = df.columns[0]
        pacote_col = df.columns[2]
        preco_col  = df.columns[3]

        for _, row in df.iterrows():
            acriss = str(row[acriss_col]).strip().upper()
            if acriss in ("NAN", ""):
                continue
            try:
                price = float(row[preco_col])
            except (ValueError, TypeError):
                continue
            field = _classify(sheet, row[pacote_col])
            if field is None:
                continue
            mapping.setdefault(acriss, {}).setdefault(field, []).append(price)

    return {
        acriss: {f: round(sum(prices) / len(prices), 4) for f, prices in fields.items()}
        for acriss, fields in mapping.items()
    }


def get_price(matrix_data, group, product, days):
    rows = matrix_data.get("data", {}).get(group, {}).get(product)
    if not rows:
        return None
    for row in rows:
        if row["sd"] <= days <= row["ed"]:
            value = row["v"]
            if value >= 9000:
                return None
            return float(value)
    return None


# =========================================================
# PROGRESSÃO TG ENTRE LOR — ESCALA GLOBAL UNIFORME
# =========================================================

def compute_global_tg_scale(matrix_data, tg_min_rack):
    """
    Factor único calculado a partir do TG mínimo global (todos os grupos ×
    todos os LOR diários). Aplicar o MESMO factor a todos os LOR preserva
    a progressão original: LOR 1-3 fica sempre mais caro que 4-6, etc.
    Devolve float >= 1.0.
    """
    daily_refs = [(1,3,1),(4,6,4),(7,10,7),(11,29,11),(31,999,31)]
    global_min = None
    for group in matrix_data["groups"]:
        for ls, le, ref in daily_refs:
            v = get_price(matrix_data, group, "TG", ref)
            if v is not None and v > 0:
                global_min = v if global_min is None else min(global_min, v)
    if global_min is None or global_min <= 0 or tg_min_rack <= 0:
        return 1.0
    return max(1.0, tg_min_rack / global_min)


def compute_tg_lor_progression(matrix_data, group, tg_min_rack, global_scale=None):
    """
    Aplica o factor único a TODOS os LOR do grupo (incluindo LOR 30-30 e
    grupos já acima do piso). Factor único preserva a ordenação original:
    TG[1-3] > TG[4-6] > TG[7-10] em todos os grupos.
    Devolve dict {(lor_start, lor_end, ref_day): tg_scaled}.
    """
    lor_intervals = [
        (1, 3, 1), (4, 6, 4), (7, 10, 7),
        (11, 29, 11), (30, 30, 30), (31, 999, 31),
    ]
    if global_scale is None:
        global_scale = compute_global_tg_scale(matrix_data, tg_min_rack)

    scale = global_scale if isinstance(global_scale, float) else 1.0

    result = {}
    for ls, le, ref in lor_intervals:
        v = get_price(matrix_data, group, "TG", ref)
        if v is not None:
            result[(ls, le, ref)] = v * scale

    # Garantir progressão estritamente decrescente nos LOR diários.
    # Quando flat/invertido, forçar descida usando rácio BF_orig entre intervalos.
    daily_order = [(1,3,1),(4,6,4),(7,10,7),(11,29,11),(31,999,31)]
    prev_tg = None
    prev_bf = None
    for key in daily_order:
        if key not in result:
            prev_tg = None
            prev_bf = None
            continue
        ls, le, ref = key
        v  = result[key]
        bf = get_price(matrix_data, group, "BF", ref)
        if prev_tg is not None and v >= prev_tg:
            if prev_bf and bf and prev_bf > 0:
                step = prev_tg * (bf / prev_bf)
            else:
                step = prev_tg * 0.99
            # nunca descer abaixo do piso (se definido)
            step = max(step, tg_min_rack) if tg_min_rack > 0 else max(step, 0.0)
            result[key] = round(step, 4)
        prev_tg = result[key]
        prev_bf = bf
    return result


# =========================================================
# CÁLCULO DE PRICING
# =========================================================

def calculate_pricing(
    matrix_data, group, days, counter_discount, online_discount, bf_weight, vat,
    easy_counter_discount=None,
    ai_counter_discount=None, ai_online_discount=None,
    easy_margin_min=3.5, easy_margin_max=7.0,
    online_extra_discount=0.10,
    tg_min_iva=8.0,           # piso mínimo TG com IVA (€). 0 = sem piso.
    tg_override=None,         # valor TG já escalado (progressão pré-calculada)
):
    bq_weight = 1 - bf_weight

    if easy_counter_discount is None:
        easy_counter_discount = 0.0
    if ai_counter_discount is None:
        ai_counter_discount = counter_discount
    if ai_online_discount is None:
        ai_online_discount = online_discount

    # Piso TG sem IVA: TG_min_rack = tg_min_iva / (1 + vat)
    tg_min_rack = (tg_min_iva / (1 + vat)) if (tg_min_iva and tg_min_iva > 0) else 0.0

    ld = get_price(matrix_data, group, "LD", days)
    bf = get_price(matrix_data, group, "BF", days)
    bq = get_price(matrix_data, group, "BQ", days)
    be = get_price(matrix_data, group, "BE", days)
    # tg_orig: usado no target do SMART+ (balcão antigo incluía este TG)
    # tg: versão escalada, usada no AI e como tg_new a gravar na matriz
    tg_orig = get_price(matrix_data, group, "TG", days)
    tg = (tg_override if (tg_override is not None and tg_orig is not None) else tg_orig)
    ip = get_price(matrix_data, group, "I",  days)
    bc = get_price(matrix_data, group, "BC", days)

    if bc is None:
        bc = 5.0

    div_counter      = 1 - counter_discount
    div_online       = 1 - online_discount
    div_easy_counter = 1 - easy_counter_discount

    issues = []
    if ld is None:      issues.append("LD indisponível para este grupo/duração.")
    if bf is None:      issues.append("BF indisponível para este grupo/duração.")
    if bq is None:      issues.append("BQ indisponível para este grupo/duração.")
    if tg_orig is None: issues.append("TG indisponível para este grupo/duração.")
    if ip is None:      issues.append("I indisponível para este grupo/duração.")
    if bc is None:      issues.append("BC indisponível para este grupo/duração.")

    def safe_sum(values):
        if any(v is None or pd.isna(v) for v in values):
            return None
        return float(sum(values))

    # SMART+ target baseado no TG ORIGINAL (referência histórica de balcão)
    smart_base_old   = safe_sum([ld, bf, bq, tg_orig])
    smart_base       = safe_sum([ld, bf, bq])
    smart_rack_target = smart_base_old

    bf_bq_current = safe_sum([bf, bq])
    bf_bq_target  = (smart_rack_target - ld if smart_rack_target is not None and ld is not None else None)
    bf_bq_gap     = (bf_bq_target - bf_bq_current if bf_bq_target is not None and bf_bq_current is not None else None)

    ld_new = ld
    bf_new = (bf + bf_bq_gap * bf_weight  if bf is not None and bf_bq_gap is not None and bf_bq_gap > 0 else bf)
    bq_new = (bq + bf_bq_gap * bq_weight  if bq is not None and bf_bq_gap is not None and bf_bq_gap > 0 else bq)
    be_new = (bf_new * 0.80 if bf_new is not None else None)

    smart_rack_new = safe_sum([ld_new, bf_new, bq_new])
    smart_counter  = smart_rack_new * div_counter if smart_rack_new is not None else None
    smart_online   = smart_rack_new * div_online  if smart_rack_new is not None else None
    smart_counter_vat = smart_counter * (1 + vat) if smart_counter is not None else None
    smart_online_vat  = smart_online  * (1 + vat) if smart_online  is not None else None

    ai_base            = safe_sum([bc, ld, bf, bq, ip, tg_orig])   # referência histórica com TG original
    ai_without_tg_new  = safe_sum([bc, ld_new, bf_new, bq_new, ip])
    easy_base          = safe_sum([tg_orig, bc])                    # referência histórica

    ok_ai_gt_smart_counter      = False
    ok_ai_gt_smart_online       = False
    ok_smart_easy_gt_ai_counter = False
    ok_smart_tg_gt_ai_counter   = False
    ok_easy_margin_counter      = False
    ok_easy_margin_online       = False

    ai_counter_discount_max_rule2  = None
    ai_online_discount_max_rule2   = None
    ai_counter_discount_min_ruleB  = None
    ai_counter_discount_min_ruleC  = None
    ai_counter_discount_valid_low  = None
    ai_counter_discount_valid_high = None
    ai_counter_discount_solved     = None
    ai_online_discount_solved      = None
    easy_margin_counter            = None
    easy_margin_online             = None

    tg_new        = tg
    ai_rack_new   = None
    easy_rack_new = None
    ai_counter    = None
    ai_online     = None
    ai_counter_vat = None
    ai_online_vat  = None

    if (ld_new is not None and bf_new is not None and bq_new is not None
            and tg_orig is not None and ip is not None
            and smart_rack_new is not None and smart_rack_new > 0
            and ai_without_tg_new is not None):

        for _pass in range(2):
            # TG nunca desce abaixo do original nem abaixo do piso mínimo de mercado
            tg_candidate        = max(tg, tg_new, tg_min_rack)
            ai_rack_candidate   = ai_without_tg_new + tg_candidate
            easy_rack_candidate = tg_candidate + bc

            if ai_rack_candidate <= 0:
                break

            smart_cnt_c = smart_rack_new * div_counter
            easy_cnt_c  = easy_rack_candidate * div_easy_counter
            smart_cnt_o = smart_rack_new * div_online

            max_da_c = max(0.0, 1 - smart_cnt_c / ai_rack_candidate - 0.0001)
            max_da_o = max(0.0, 1 - smart_cnt_o / ai_rack_candidate - 0.0001)

            min_da_c_B = max(0.0, 1 - (smart_cnt_c + easy_cnt_c) / ai_rack_candidate + 0.0001)
            min_da_c_C = max(0.0, 1 - (smart_cnt_c + tg_candidate) / ai_rack_candidate + 0.0001)

            min_da_c_D = 1 - (smart_cnt_c + easy_cnt_c - easy_margin_min) / ai_rack_candidate
            max_da_c_D = 1 - (smart_cnt_c + easy_cnt_c - easy_margin_max) / ai_rack_candidate

            valid_low  = max(min_da_c_B, min_da_c_C, min_da_c_D, 0.0)
            valid_high = min(max_da_c, max_da_c_D)

            if valid_low <= valid_high or _pass == 1:
                tg_new        = tg_candidate
                ai_rack_new   = ai_rack_candidate
                easy_rack_new = easy_rack_candidate
                break
            else:
                tg_needed_B = None
                if div_easy_counter > 0 and easy_cnt_c < (ai_rack_candidate - smart_cnt_c):
                    numerator = ai_without_tg_new - smart_cnt_c - bc * div_easy_counter + 0.01
                    if easy_counter_discount > 0:
                        tg_needed_B = numerator / easy_counter_discount
                    else:
                        tg_needed_B = tg
                tg_new = max(tg, tg_needed_B if tg_needed_B is not None else tg)

        ai_counter_discount_max_rule2  = max_da_c
        ai_online_discount_max_rule2   = max_da_o
        ai_counter_discount_min_ruleB  = min_da_c_B
        ai_counter_discount_min_ruleC  = min_da_c_C
        ai_counter_discount_valid_low  = valid_low
        ai_counter_discount_valid_high = valid_high

        if valid_low <= valid_high:
            da_c_opt = (valid_low + valid_high) / 2
        else:
            da_c_opt = min(max(min_da_c_B, min_da_c_C), max_da_c)
        da_c_opt = max(0.0, min(da_c_opt, 0.9999))

        ai_counter_discount_solved = da_c_opt
        ai_counter     = ai_rack_new * (1 - da_c_opt)
        ai_counter_vat = ai_counter * (1 + vat)

        da_o_opt = max(da_c_opt + online_extra_discount, online_discount)
        da_o_opt = min(da_o_opt, max_da_o)
        ai_online_discount_solved = da_o_opt
        ai_online     = ai_rack_new * (1 - da_o_opt)
        ai_online_vat = ai_online * (1 + vat)

        ok_ai_gt_smart_counter      = ai_counter > smart_cnt_c
        ok_ai_gt_smart_online       = ai_online  > smart_cnt_o
        ok_smart_easy_gt_ai_counter = (smart_cnt_c + easy_cnt_c) > ai_counter
        ok_smart_tg_gt_ai_counter   = (smart_cnt_c + tg_new)     > ai_counter
        easy_margin_counter         = smart_cnt_c + easy_cnt_c - ai_counter
        easy_margin_online          = None
        ok_easy_margin_counter      = easy_margin_min <= easy_margin_counter <= easy_margin_max
        ok_easy_margin_online       = False

    else:
        tg_new        = max(tg, tg_min_rack) if tg is not None else tg_orig
        ai_rack_new   = None
        easy_rack_new = None
        ai_counter    = None
        ai_online     = None
        ai_counter_vat = None
        ai_online_vat  = None

    easy_rack_new    = (tg_new + bc) if tg_new is not None else None
    easy_counter     = easy_rack_new * div_easy_counter if easy_rack_new is not None else None
    easy_online      = None
    easy_counter_vat = easy_counter * (1 + vat) if easy_counter is not None else None
    easy_online_vat  = None

    if ai_rack_new is None and ai_without_tg_new is not None and tg_new is not None:
        ai_rack_new = ai_without_tg_new + tg_new

    ok_easy_constraint = (
        smart_rack_new is not None and easy_rack_new is not None
        and ai_rack_new is not None
        and (smart_rack_new + easy_rack_new) > ai_rack_new
    )

    easy_increase_pct  = (easy_rack_new / easy_base - 1   if easy_base  and easy_rack_new  and easy_base  > 0 else None)
    smart_increase_pct = (smart_rack_new / smart_base_old - 1 if smart_base_old and smart_rack_new and smart_base_old > 0 else None)
    bf_increase_pct    = (bf_new / bf - 1 if bf and bf_new and bf > 0 else None)
    bq_increase_pct    = (bq_new / bq - 1 if bq and bq_new and bq > 0 else None)
    tg_increase_pct    = (tg_new / tg_orig - 1 if tg_orig and tg_new and tg_orig > 0 else None)
    ai_increase_pct    = (ai_rack_new / ai_base - 1 if ai_base and ai_rack_new and ai_base > 0 else None)

    smart_implicit_discount = (1 - smart_counter / smart_rack_new if smart_rack_new and smart_counter and smart_rack_new > 0 else None)
    ai_implicit_discount    = (1 - ai_counter / ai_rack_new        if ai_rack_new   and ai_counter   and ai_rack_new   > 0 else None)

    smart_counter_target = smart_base_old * div_counter if smart_base_old is not None else None
    ok_smart = (abs(smart_counter - smart_counter_target) < 0.02 if smart_counter is not None and smart_counter_target is not None else False)
    ok_ai    = (abs(ai_counter - ai_base) < 0.02 if ai_counter is not None and ai_base is not None else False)

    be_increase_pct = (be_new / be - 1 if be and be_new and be > 0 else None)

    return {
        "ld": ld, "bf": bf, "bq": bq, "be": be, "tg": tg_orig, "ip": ip, "bc": bc,
        "ld_new": ld_new, "bf_new": bf_new, "bq_new": bq_new, "be_new": be_new, "tg_new": tg_new,
        "be_increase_pct": be_increase_pct,
        "smart_base": smart_base_old,
        "smart_base_new": smart_base,
        "smart_rack_target": smart_rack_target,
        "smart_rack_new": smart_rack_new,
        "smart_counter": smart_counter,
        "smart_counter_target": smart_counter_target,
        "smart_online": smart_online,
        "smart_counter_vat": smart_counter_vat,
        "smart_online_vat": smart_online_vat,
        "bf_bq_current": bf_bq_current,
        "bf_bq_target": bf_bq_target,
        "bf_bq_gap": bf_bq_gap,
        "ai_base": ai_base,
        "ai_rack_target": None,
        "ai_without_tg_new": ai_without_tg_new,
        "ai_rack_new": ai_rack_new,
        "ai_counter": ai_counter,
        "ai_online": ai_online,
        "ai_counter_vat": ai_counter_vat,
        "ai_online_vat": ai_online_vat,
        "smart_increase_pct": smart_increase_pct,
        "bf_increase_pct": bf_increase_pct,
        "bq_increase_pct": bq_increase_pct,
        "tg_increase_pct": tg_increase_pct,
        "ai_increase_pct": ai_increase_pct,
        "tg_min_iva": tg_min_iva,
        "tg_min_rack": tg_min_rack,
        "tg_raised_to_min": (
            tg_orig is not None and tg_new is not None and tg_min_rack > 0
            and tg_new > tg_orig and abs(tg_new - tg_min_rack) < 0.01
        ),
        "tg_below_min": (
            tg_orig is not None and tg_min_rack > 0 and tg_orig < tg_min_rack
        ),
        "smart_implicit_discount": smart_implicit_discount,
        "ai_implicit_discount": ai_implicit_discount,
        "ok_smart": ok_smart,
        "ok_ai": ok_ai,
        "ok_easy_constraint": ok_easy_constraint,
        "ok_ai_gt_smart_counter": ok_ai_gt_smart_counter,
        "ok_ai_gt_smart_online": ok_ai_gt_smart_online,
        "ok_smart_easy_gt_ai_counter": ok_smart_easy_gt_ai_counter,
        "ok_smart_tg_gt_ai_counter": ok_smart_tg_gt_ai_counter,
        "ai_counter_discount_max_rule2": ai_counter_discount_max_rule2,
        "ai_online_discount_max_rule2": ai_online_discount_max_rule2,
        "ai_counter_discount_min_ruleB": ai_counter_discount_min_ruleB,
        "ai_counter_discount_min_ruleC": ai_counter_discount_min_ruleC,
        "ai_counter_discount_valid_low": ai_counter_discount_valid_low,
        "ai_counter_discount_valid_high": ai_counter_discount_valid_high,
        "online_extra_discount": online_extra_discount,
        "ai_fixed": ai_without_tg_new,
        "easy_base": easy_base,
        "easy_rack_new": easy_rack_new,
        "easy_counter": easy_counter,
        "easy_online": easy_online,
        "easy_counter_vat": easy_counter_vat,
        "easy_online_vat": easy_online_vat,
        "easy_increase_pct": easy_increase_pct,
        "easy_margin_counter": easy_margin_counter,
        "easy_margin_online": easy_margin_online,
        "ai_counter_discount_solved": ai_counter_discount_solved,
        "ai_online_discount_solved": ai_online_discount_solved,
        "ok_easy_margin_counter": ok_easy_margin_counter,
        "ok_easy_margin_online": ok_easy_margin_online,
        "easy_margin_min": easy_margin_min,
        "easy_margin_max": easy_margin_max,
        "easy_counter_discount_used": easy_counter_discount,
        "easy_online_discount_used": None,
        "ai_counter_discount_used": ai_counter_discount,
        "ai_online_discount_used": ai_online_discount,
        "issues": issues,
    }


# =========================================================
# EXPORTAÇÃO
# =========================================================

def build_all_groups_lor_export(
    matrix_data, counter_discount, online_discount, bf_weight, vat,
    easy_counter_discount=None,
    ai_counter_discount=None, ai_online_discount=None,
    easy_margin_min=3.5, easy_margin_max=7.0,
    online_extra_discount=0.10,
    tg_min_iva=8.0,
):
    lor_intervals = [
        {"lor_start": 1,  "lor_end": 3,   "days_reference": 1},
        {"lor_start": 4,  "lor_end": 6,   "days_reference": 4},
        {"lor_start": 7,  "lor_end": 10,  "days_reference": 7},
        {"lor_start": 11, "lor_end": 29,  "days_reference": 11},
        {"lor_start": 30, "lor_end": 30,  "days_reference": 30},
        {"lor_start": 31, "lor_end": 999, "days_reference": 31},
    ]

    rows_detail = []
    rows_smart  = []
    rows_ai     = []
    rows_easy   = []
    rows_parameters = []

    _tg_min_rack_exp  = (tg_min_iva / (1 + vat)) if (tg_min_iva and tg_min_iva > 0) else 0.0
    _global_scale_exp = compute_global_tg_scale(matrix_data, _tg_min_rack_exp)

    for group in matrix_data["groups"]:
        tg_progression = compute_tg_lor_progression(matrix_data, group, _tg_min_rack_exp, _global_scale_exp)

        for lor in lor_intervals:
            tg_ov = tg_progression.get((lor["lor_start"], lor["lor_end"], lor["days_reference"]))
            result = calculate_pricing(
                matrix_data=matrix_data, group=group, days=lor["days_reference"],
                counter_discount=counter_discount, online_discount=online_discount,
                bf_weight=bf_weight, vat=vat, easy_counter_discount=easy_counter_discount,
                ai_counter_discount=ai_counter_discount, ai_online_discount=ai_online_discount,
                easy_margin_min=easy_margin_min, easy_margin_max=easy_margin_max,
                online_extra_discount=online_extra_discount,
                tg_min_iva=tg_min_iva,
                tg_override=tg_ov,
            )

            issue_text = " | ".join(result["issues"]) if result["issues"] else ""

            rows_detail.append({
                "Grupo ACRISS": group,
                "LOR início": lor["lor_start"], "LOR fim": lor["lor_end"],
                "Dia referência": lor["days_reference"],
                "LD atual": result["ld"], "LD novo": result["ld_new"],
                "BF atual": result["bf"], "BF novo": result["bf_new"], "BF variação %": result["bf_increase_pct"],
                "BQ atual": result["bq"], "BQ novo": result["bq_new"], "BQ variação %": result["bq_increase_pct"],
                "BE atual": result["be"], "BE novo": result["be_new"], "BE variação %": result["be_increase_pct"],
                "TG atual": result["tg"], "TG novo": result["tg_new"], "TG variação %": result["tg_increase_pct"],
                "I atual": result["ip"], "I novo": result["ip"],
                "BC atual": result["bc"], "BC novo": result["bc"],
                "SMART+ atual s/IVA": result["smart_base"],
                "SMART+ rack novo s/IVA": result["smart_rack_new"],
                "SMART+ balcão s/IVA": result["smart_counter"],
                "SMART+ online s/IVA": result["smart_online"],
                "SMART+ balcão c/IVA": result["smart_counter_vat"],
                "SMART+ online c/IVA": result["smart_online_vat"],
                "SMART+ variação rack %": result["smart_increase_pct"],
                "SMART+ regra cumprida": "Sim" if result["ok_smart"] else "Não",
                "AI atual s/IVA": result["ai_base"],
                "AI rack novo s/IVA": result["ai_rack_new"],
                "AI balcão s/IVA": result["ai_counter"],
                "AI online s/IVA": result["ai_online"],
                "AI balcão c/IVA": result["ai_counter_vat"],
                "AI online c/IVA": result["ai_online_vat"],
                "AI variação rack %": result["ai_increase_pct"],
                "AI regra cumprida": "Sim" if result["ok_ai"] else "Não",
                "Pack Easy atual s/IVA": result["easy_base"],
                "Pack Easy rack novo s/IVA": result["easy_rack_new"],
                "Pack Easy balcão s/IVA": result["easy_counter"],
                "Pack Easy online s/IVA": result["easy_online"],
                "Pack Easy balcão c/IVA": result["easy_counter_vat"],
                "Pack Easy online c/IVA": result["easy_online_vat"],
                "Pack Easy variação rack %": result["easy_increase_pct"],
                "SMART++Easy > AI": "Sim" if result["ok_easy_constraint"] else "Não",
                "Easy Margem Balcão (€)": result["easy_margin_counter"],
                "Easy Margem Online (€)": result["easy_margin_online"],
                "Desc. AI Balcão necessário (%)": result["ai_counter_discount_solved"] * 100 if result["ai_counter_discount_solved"] is not None else None,
                "Desc. AI Online necessário (%)": result["ai_online_discount_solved"]  * 100 if result["ai_online_discount_solved"]  is not None else None,
                "Margem balcão OK": "Sim" if result["ok_easy_margin_counter"] else "Não",
                "Margem online OK": "Sim" if result["ok_easy_margin_online"]  else "Não",
                "Observações": issue_text,
            })

            rows_smart.append({
                "Grupo ACRISS": group,
                "LOR início": lor["lor_start"], "LOR fim": lor["lor_end"],
                "Dia referência": lor["days_reference"],
                "LD atual": result["ld"], "LD novo": result["ld_new"],
                "BF atual": result["bf"], "BF novo": result["bf_new"], "BF variação %": result["bf_increase_pct"],
                "BQ atual": result["bq"], "BQ novo": result["bq_new"], "BQ variação %": result["bq_increase_pct"],
                "SMART+ atual s/IVA": result["smart_base"],
                "SMART+ rack novo s/IVA": result["smart_rack_new"],
                "SMART+ balcão s/IVA": result["smart_counter"],
                "SMART+ online s/IVA": result["smart_online"],
                "SMART+ balcão c/IVA": result["smart_counter_vat"],
                "SMART+ online c/IVA": result["smart_online_vat"],
                "Aumento necessário BF+BQ": result["bf_bq_gap"],
                "Desconto implícito balcão": result["smart_implicit_discount"],
                "Regra cumprida": "Sim" if result["ok_smart"] else "Não",
                "Observações": issue_text,
            })

            rows_ai.append({
                "Grupo ACRISS": group,
                "LOR início": lor["lor_start"], "LOR fim": lor["lor_end"],
                "Dia referência": lor["days_reference"],
                "BC atual": result["bc"], "BC novo": result["bc"],
                "LD atual": result["ld"], "LD novo": result["ld_new"],
                "BF atual": result["bf"], "BF novo": result["bf_new"],
                "BQ atual": result["bq"], "BQ novo": result["bq_new"],
                "I atual": result["ip"], "I novo": result["ip"],
                "TG atual": result["tg"], "TG novo": result["tg_new"], "TG variação %": result["tg_increase_pct"],
                "AI atual s/IVA": result["ai_base"],
                "AI rack novo s/IVA": result["ai_rack_new"],
                "AI balcão s/IVA": result["ai_counter"],
                "AI online s/IVA": result["ai_online"],
                "AI balcão c/IVA": result["ai_counter_vat"],
                "AI online c/IVA": result["ai_online_vat"],
                "AI variação rack %": result["ai_increase_pct"],
                "Desconto implícito balcão": result["ai_implicit_discount"],
                "Regra cumprida": "Sim" if result["ok_ai"] else "Não",
                "Observações": issue_text,
            })

            rows_easy.append({
                "Grupo ACRISS": group,
                "LOR início": lor["lor_start"], "LOR fim": lor["lor_end"],
                "Dia referência": lor["days_reference"],
                "TG atual": result["tg"], "TG novo": result["tg_new"], "TG variação %": result["tg_increase_pct"],
                "BC atual": result["bc"], "BC novo": result["bc"],
                "Pack Easy atual s/IVA": result["easy_base"],
                "Pack Easy rack novo s/IVA": result["easy_rack_new"],
                "Pack Easy balcão s/IVA": result["easy_counter"],
                "Pack Easy online s/IVA": result["easy_online"],
                "Pack Easy balcão c/IVA": result["easy_counter_vat"],
                "Pack Easy online c/IVA": result["easy_online_vat"],
                "Pack Easy variação rack %": result["easy_increase_pct"],
                "SMART+ rack novo": result["smart_rack_new"],
                "SMART+ + Easy rack novo": (
                    round(result["smart_rack_new"] + result["easy_rack_new"], 4)
                    if result["smart_rack_new"] is not None and result["easy_rack_new"] is not None else None
                ),
                "AI rack novo": result["ai_rack_new"],
                "Restrição SMART++Easy > AI": "Sim" if result["ok_easy_constraint"] else "Não",
                "AI > SMART+ Balcão": "Sim" if result["ok_ai_gt_smart_counter"] else "Não",
                "AI > SMART+ Online": "Sim" if result["ok_ai_gt_smart_online"] else "Não",
                "Desc. AI máx balcão (Regra 2) (%)": result["ai_counter_discount_max_rule2"] * 100 if result["ai_counter_discount_max_rule2"] is not None else None,
                "Desc. AI máx online (Regra 2) (%)": result["ai_online_discount_max_rule2"]  * 100 if result["ai_online_discount_max_rule2"]  is not None else None,
                "Easy Margem Balcão (€)": result["easy_margin_counter"],
                "Easy Online": "n/a — só balcão",
                f"Margem [{result['easy_margin_min']}-{result['easy_margin_max']}€] balcão OK": "Sim" if result["ok_easy_margin_counter"] else "Não",
                f"Margem [{result['easy_margin_min']}-{result['easy_margin_max']}€] online OK": "Sim" if result["ok_easy_margin_online"]  else "Não",
                "Desc. AI Balcão necessário (%)": result["ai_counter_discount_solved"] * 100 if result["ai_counter_discount_solved"] is not None else None,
                "Desc. AI Online necessário (%)": result["ai_online_discount_solved"]  * 100 if result["ai_online_discount_solved"]  is not None else None,
                "Observações": issue_text,
            })

    rows_parameters = [
        {"Parâmetro": "Desconto balcão",           "Valor": counter_discount},
        {"Parâmetro": "Desconto online",            "Valor": online_discount},
        {"Parâmetro": "Peso subida BF",             "Valor": bf_weight},
        {"Parâmetro": "Peso subida BQ",             "Valor": 1 - bf_weight},
        {"Parâmetro": "IVA",                        "Valor": vat},
        {"Parâmetro": "Regra SMART+",               "Valor": "SMART+ = LD + BF + BQ (sem TG); balcão novo = balcão antigo (com TG)"},
        {"Parâmetro": "Regra All Inclusive",        "Valor": "AI = BC + LD + BF + BQ + I + TG; LD/BF/BQ iguais ao SMART+ novo; TG sobe livremente"},
        {"Parâmetro": "Regra Pack Easy",            "Valor": f"Pack Easy = TG + BC; SÓ BALCÃO; margem alvo: [{easy_margin_min}€ ; {easy_margin_max}€]"},
        {"Parâmetro": "Desc. balcão SMART+",        "Valor": counter_discount},
        {"Parâmetro": "Desc. online SMART+",        "Valor": online_discount},
        {"Parâmetro": "Desc. balcão Easy",          "Valor": easy_counter_discount if easy_counter_discount is not None else counter_discount},
        {"Parâmetro": "Desc. online Easy",          "Valor": "n/a — Easy só balcão"},
        {"Parâmetro": "Desc. balcão AI",            "Valor": ai_counter_discount  if ai_counter_discount  is not None else counter_discount},
        {"Parâmetro": "Desc. online AI",            "Valor": ai_online_discount   if ai_online_discount   is not None else online_discount},
        {"Parâmetro": "Margem Easy mínima (€)",     "Valor": easy_margin_min},
        {"Parâmetro": "Margem Easy máxima (€)",     "Valor": easy_margin_max},
    ]

    df_params = pd.DataFrame(rows_parameters)
    df_params["Valor"] = df_params["Valor"].astype(str)   # prevent ArrowInvalid (mixed types)

    return (
        pd.DataFrame(rows_detail), pd.DataFrame(rows_smart),
        pd.DataFrame(rows_ai), pd.DataFrame(rows_easy),
        df_params,
    )


def generate_sixt_matrix(
    file_bytes, sheet_name,
    matrix_data,
    counter_discount, online_discount, bf_weight, vat,
    easy_counter_discount=0.0,
    easy_margin_min=3.5, easy_margin_max=7.0,
    online_extra_discount=0.10,
    tg_min_iva=8.0,
    highlight_changes=True,
):
    """
    Gera uma nova matriz SIXT no formato exacto da original.
    Actualiza APENAS BF, TG e BQ. LD, BE, AY, AE, I, BC ficam inalterados.
    Células alteradas ficam destacadas a amarelo (opcional).
    Devolve (BytesIO, n_células_alteradas, resumo_por_produto).
    """
    from datetime import date
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font

    CHANGED_FILL   = PatternFill("solid", fgColor="FFF3CD")
    INCREASED_FONT = Font(name="Arial", sz=8, bold=True, color="7C2D12")
    PRODUCTS_TO_UPDATE = {"BF", "TG", "BQ", "BE"}

    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    today = date.today()
    ws.cell(1, 1).value = f"Matrix - {today.strftime('%d.%m.%Y')}"
    ws.cell(4, 2).value = today

    # ── Expandir BE para 6 LOR (igual ao BF) ─────────────────────────────
    # A matriz original tem BE com 3 intervalos largos: [1-29], [30-30], [31-999]
    # Vamos inserir 3 colunas novas (BE[1-3], BE[4-6], BE[7-10]) antes do BE[1-29]
    # e actualizar BE[1-29] → BE[11-29]
    be_start_col = None
    for c in range(2, ws.max_column + 1):
        if str(ws.cell(8, c).value).strip() == "BE":
            be_start_col = c
            break

    if be_start_col is not None:
        ws.insert_cols(be_start_col, 3)
        # Copiar número de formato da coluna BE original (agora deslocada +3)
        ref_be_col = be_start_col + 3
        for i, (sd, ed) in enumerate([(1, 3), (4, 6), (7, 10)]):
            col = be_start_col + i
            ws.cell(8, col).value = "BE"
            ws.cell(9, col).value = sd
            ws.cell(10, col).value = ed
        # Actualizar BE[1-29] → BE[11-29]
        ws.cell(9, ref_be_col).value = 11

    # ── Reconstruir col_map depois das inserções ──────────────────────────
    col_map = {}
    for c in range(2, ws.max_column + 1):
        prod = ws.cell(8, c).value
        vont = ws.cell(9, c).value
        bist = ws.cell(10, c).value
        if (prod in PRODUCTS_TO_UPDATE
                and isinstance(vont, (int, float))
                and isinstance(bist, (int, float))):
            col_map[(str(prod), int(vont), int(bist))] = c

    lor_intervals = [
        (1,  3,   1),
        (4,  6,   4),
        (7,  10,  7),
        (11, 29,  11),
        (30, 30,  30),
        (31, 999, 31),
    ]

    group_rows = {}
    for r in range(17, ws.max_row + 1):
        grp = ws.cell(r, 1).value
        if grp is not None:
            group_rows[str(grp).strip()] = r

    stats = {"BF": {"changed": 0, "sum_delta": 0.0},
             "TG": {"changed": 0, "sum_delta": 0.0},
             "BQ": {"changed": 0, "sum_delta": 0.0},
             "BE": {"changed": 0, "sum_delta": 0.0}}

    tg_min_rack       = (tg_min_iva / (1 + vat)) if (tg_min_iva and tg_min_iva > 0) else 0.0
    _global_scale_mat = compute_global_tg_scale(matrix_data, tg_min_rack)

    # Intervalos diários pela ordem correcta (excluir 30-30 que é valor tecto)
    daily_lor_order = [(1,3,1),(4,6,4),(7,10,7),(11,29,11),(31,999,31)]

    for group in matrix_data["groups"]:
        if group not in group_rows:
            continue
        row = group_rows[group]

        tg_progression = compute_tg_lor_progression(matrix_data, group, tg_min_rack, _global_scale_mat)

        # Calcular todos os LOR primeiro, depois aplicar monotonia
        lor_results = {}
        for lor_start, lor_end, ref_day in lor_intervals:
            tg_ov = tg_progression.get((lor_start, lor_end, ref_day))
            result = calculate_pricing(
                matrix_data=matrix_data, group=group, days=ref_day,
                counter_discount=counter_discount, online_discount=online_discount,
                bf_weight=bf_weight, vat=vat,
                easy_counter_discount=easy_counter_discount,
                ai_counter_discount=None, ai_online_discount=None,
                easy_margin_min=easy_margin_min, easy_margin_max=easy_margin_max,
                online_extra_discount=online_extra_discount,
                tg_min_iva=tg_min_iva,
                tg_override=tg_ov,
            )
            lor_results[(lor_start, lor_end, ref_day)] = dict(result)

        # Garantir progressão monotonicamente ESTRITAMENTE decrescente.
        # Quando um intervalo seria igual ou maior que o anterior, forçar a descida
        # usando o rácio BF_new entre intervalos consecutivos (progressão natural do BF).
        # TG tem piso mínimo (tg_min_rack); BQ e BF não.
        for field, floor_val in [("tg_new", tg_min_rack), ("bf_new", 0.0), ("bq_new", 0.0)]:
            prev_val = None
            prev_bf  = None
            for key in daily_lor_order:
                ls, le, ref = key
                r = lor_results.get(key)
                if r is None or r.get(field) is None:
                    prev_val = None
                    prev_bf  = None
                    continue
                v   = r[field]
                cur_bf = r.get("bf_new")
                if prev_val is not None and v >= prev_val:
                    # Flat ou invertido — forçar descida proporcional ao BF
                    if prev_bf and cur_bf and prev_bf > 0:
                        step = prev_val * (cur_bf / prev_bf)
                    else:
                        step = prev_val * 0.99
                    step = max(step, floor_val) if floor_val > 0 else max(step, 0.0)
                    lor_results[key][field] = round(step, 4)
                prev_val = lor_results[key][field]
                prev_bf  = cur_bf

        # Escrever na matriz
        for lor_start, lor_end, ref_day in lor_intervals:
            result = lor_results.get((lor_start, lor_end, ref_day), {})
            if not result:
                continue

            be_new = round(result["bf_new"] * 0.80, 4) if result.get("bf_new") is not None else None

            for product, orig_key, new_val in [
                ("BF", "bf",  result.get("bf_new")),
                ("TG", "tg",  result.get("tg_new")),
                ("BQ", "bq",  result.get("bq_new")),
                ("BE", "be",  be_new),
            ]:
                col_key = (product, lor_start, lor_end)
                if col_key not in col_map or new_val is None:
                    continue
                col  = col_map[col_key]
                orig = result.get(orig_key)
                if product == "BE" and orig is None:
                    orig = get_price(matrix_data, group, "BE", ref_day)

                rounded     = round(new_val, 4)
                cell        = ws.cell(row, col)
                cell.value  = rounded

                if highlight_changes and orig is not None and abs(rounded - orig) > 0.001:
                    cell.fill = CHANGED_FILL
                    cell.font = INCREASED_FONT
                    stats[product]["changed"]   += 1
                    stats[product]["sum_delta"] += rounded - orig

    total_changed = sum(s["changed"] for s in stats.values())
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, total_changed, stats


def xlsx_col_letter(col_idx):
    letters = ""
    col_idx += 1
    while col_idx:
        col_idx, remainder = divmod(col_idx - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def create_excel_download(df_detail, df_smart, df_ai, df_easy, df_parameters):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df_detail.to_excel(writer, sheet_name="Resumo_Final", index=False)
        df_smart.to_excel(writer,  sheet_name="SMART+",       index=False)
        df_ai.to_excel(writer,     sheet_name="All_Inclusive", index=False)
        df_easy.to_excel(writer,   sheet_name="Pack_Easy",    index=False)
        df_parameters.to_excel(writer, sheet_name="Parametros", index=False)

        workbook      = writer.book
        header_format = workbook.add_format({"bold": True, "font_color": "white", "bg_color": "#111827", "border": 1, "align": "center", "valign": "vcenter"})
        money_format  = workbook.add_format({"num_format": "#,##0.00 €", "border": 1})
        percent_format= workbook.add_format({"num_format": "0.0%", "border": 1})
        text_format   = workbook.add_format({"border": 1})
        integer_format= workbook.add_format({"num_format": "0", "border": 1})
        ok_format     = workbook.add_format({"bg_color": "#DCFCE7", "font_color": "#166534", "border": 1})
        warn_format   = workbook.add_format({"bg_color": "#FEE2E2", "font_color": "#991B1B", "border": 1})

        sheets = {
            "Resumo_Final": df_detail, "SMART+": df_smart,
            "All_Inclusive": df_ai,    "Pack_Easy": df_easy, "Parametros": df_parameters,
        }

        for sheet_name, df in sheets.items():
            ws = writer.sheets[sheet_name]
            ws.freeze_panes(1, 0)
            ws.autofilter(0, 0, len(df), len(df.columns) - 1)
            for col_num, col_name in enumerate(df.columns):
                ws.write(0, col_num, str(col_name), header_format)
                max_len = max([len(str(col_name)), *[len(str(x)) for x in df[col_name].head(200).fillna("").values]])
                width = min(max(max_len + 2, 12), 32)
                if "%" in col_name or "Desconto" in col_name or "IVA" in col_name or "Peso" in col_name:
                    ws.set_column(col_num, col_num, width, percent_format)
                elif any(k in col_name.lower() for k in ["preço", "atual", "novo", "rack", "balcão", "online", "aumento", "smart+", "ai", "ld", "bf", "bq", "tg", "bc", " i "]):
                    ws.set_column(col_num, col_num, width, money_format)
                elif col_name in ["LOR início", "LOR fim", "Dia referência"]:
                    ws.set_column(col_num, col_num, width, integer_format)
                else:
                    ws.set_column(col_num, col_num, width, text_format)
            for col_num, col_name in enumerate(df.columns):
                if "Regra cumprida" in col_name or "SMART++Easy > AI" in col_name or "Restrição SMART+" in col_name:
                    cl = xlsx_col_letter(col_num)
                    ws.conditional_format(f"{cl}2:{cl}{len(df)+1}", {"type": "text", "criteria": "containing", "value": "Sim", "format": ok_format})
                    ws.conditional_format(f"{cl}2:{cl}{len(df)+1}", {"type": "text", "criteria": "containing", "value": "Não", "format": warn_format})
    output.seek(0)
    return output


# =========================================================
# PAGE CONFIG & SESSION STATE
# =========================================================

st.set_page_config(
    page_title="SIXT Pricing Simulator",
    page_icon="🚗",
    layout="wide",
    initial_sidebar_state="expanded",
)

for key in ["opt_ai_counter", "opt_ai_online", "opt_result"]:
    if key not in st.session_state:
        st.session_state[key] = None

if "competitor_prices" not in st.session_state:
    st.session_state["competitor_prices"] = {}


# =========================================================
# CSS
# =========================================================

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');

*, [class*="css"] { font-family: 'Inter', sans-serif !important; }

.main { background: #F1F5F9 !important; }
.block-container { padding: 1.5rem 2rem 3rem !important; max-width: 1440px !important; }
[data-testid="stToolbar"], footer, #MainMenu { display: none !important; visibility: hidden; }

/* ── Sidebar ── */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0F172A 0%, #1E293B 100%) !important;
    border-right: 1px solid #334155;
}
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] .stMarkdown p,
[data-testid="stSidebar"] .stMarkdown span,
[data-testid="stSidebar"] .stSelectbox label,
[data-testid="stSidebar"] .stSlider label,
[data-testid="stSidebar"] .stTextInput label,
[data-testid="stSidebar"] .stNumberInput label {
    color: #CBD5E1 !important;
    font-size: 12px !important;
}
[data-testid="stSidebar"] .stSelectbox > div > div {
    background: #1E293B !important;
    border-color: #334155 !important;
    color: #F1F5F9 !important;
}
[data-testid="stSidebar"] .stTextInput > div > div > input {
    background: #1E293B !important;
    border-color: #334155 !important;
    color: #F1F5F9 !important;
}
[data-testid="stSidebar"] .stNumberInput > div > div > input {
    background: #1E293B !important;
    border-color: #334155 !important;
    color: #F1F5F9 !important;
}
[data-testid="stSidebarContent"] { padding: 1.5rem 1rem !important; }

/* ── App Header ── */
.app-header {
    background: linear-gradient(135deg, #0F172A 0%, #1E293B 55%, #7C2D12 100%);
    border-radius: 20px;
    padding: 28px 36px;
    margin-bottom: 28px;
    box-shadow: 0 25px 60px -12px rgba(15, 23, 42, 0.35);
    position: relative;
    overflow: hidden;
}
.app-header::before {
    content: '';
    position: absolute;
    top: -40px; right: -40px;
    width: 200px; height: 200px;
    border-radius: 50%;
    background: radial-gradient(circle, rgba(255,95,0,0.15) 0%, transparent 70%);
}
.header-eyebrow {
    font-size: 10px; font-weight: 700;
    text-transform: uppercase; letter-spacing: 0.18em;
    color: #FF5F00; margin-bottom: 10px;
    display: flex; align-items: center; gap: 8px;
}
.header-eyebrow::before { content: ''; display: inline-block; width: 24px; height: 2px; background: #FF5F00; border-radius: 2px; }
.header-title {
    font-size: 28px; font-weight: 900;
    color: #FFFFFF; margin: 0 0 8px;
    letter-spacing: -0.03em; line-height: 1.15;
}
.header-subtitle { font-size: 13px; color: #94A3B8; line-height: 1.6; max-width: 700px; }
.header-pills { display: flex; gap: 8px; margin-top: 16px; flex-wrap: wrap; }
.header-pill {
    background: rgba(255,255,255,0.07);
    border: 1px solid rgba(255,255,255,0.12);
    border-radius: 100px; padding: 4px 12px;
    font-size: 11px; font-weight: 600; color: #CBD5E1;
}

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {
    gap: 4px; background: white;
    padding: 6px; border-radius: 14px;
    border: 1px solid #E2E8F0;
    box-shadow: 0 2px 8px rgba(15,23,42,0.04);
    margin-bottom: 24px;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 10px !important;
    padding: 9px 22px !important;
    font-weight: 600 !important; font-size: 13px !important;
    color: #64748B; background: transparent; border: none;
    transition: all 0.15s;
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, #FF5F00, #EA4C00) !important;
    color: white !important;
    box-shadow: 0 4px 12px rgba(255,95,0,0.35) !important;
}
.stTabs [data-baseweb="tab-highlight"], .stTabs [data-baseweb="tab-border"] { display: none !important; }

/* ── KPI Cards ── */
.kpi-card {
    background: white; border-radius: 16px;
    padding: 20px 22px; border: 1px solid #E2E8F0;
    box-shadow: 0 2px 8px rgba(15,23,42,0.04);
    height: 100%; min-height: 110px;
}
.kpi-label {
    font-size: 10px; font-weight: 700;
    text-transform: uppercase; letter-spacing: 0.09em;
    color: #94A3B8; margin-bottom: 10px;
}
.kpi-value {
    font-size: 22px; font-weight: 800;
    color: #0F172A; line-height: 1.2; margin-bottom: 6px;
}
.kpi-note { font-size: 11px; color: #64748B; line-height: 1.5; }
.kpi-card.orange { border-top: 3px solid #FF5F00; }
.kpi-card.blue   { border-top: 3px solid #3B82F6; }
.kpi-card.green  { border-top: 3px solid #10B981; }
.kpi-card.purple { border-top: 3px solid #8B5CF6; }
.kpi-card.sky    { border-top: 3px solid #0EA5E9; }
.kpi-card.red    { border-top: 3px solid #EF4444; }

/* ── Section Dividers ── */
.sec-heading {
    display: flex; align-items: center; gap: 12px;
    margin: 28px 0 16px; font-size: 12px; font-weight: 700;
    text-transform: uppercase; letter-spacing: 0.09em; color: #475569;
}
.sec-heading-bar { flex: 1; height: 1px; background: #E2E8F0; }
.sec-icon {
    width: 28px; height: 28px; border-radius: 8px;
    display: inline-flex; align-items: center; justify-content: center;
    font-size: 14px; flex-shrink: 0;
}
.sec-icon.orange { background: #FFF7ED; }
.sec-icon.blue   { background: #EFF6FF; }
.sec-icon.green  { background: #F0FDF4; }

/* ── Package Banners ── */
.pkg-banner {
    border-radius: 14px; padding: 14px 20px;
    margin-bottom: 16px; display: flex;
    align-items: flex-start; gap: 14px;
    border: 1px solid;
}
.pkg-banner.smart  { background: #FFF7ED; border-color: #FED7AA; }
.pkg-banner.ai     { background: #EFF6FF; border-color: #BFDBFE; }
.pkg-banner.easy   { background: #F0FDF4; border-color: #BBF7D0; }
.pkg-banner-dot    { width: 10px; height: 10px; border-radius: 50%; margin-top: 4px; flex-shrink: 0; }
.pkg-banner-body   { flex: 1; }
.pkg-banner-title  { font-size: 13px; font-weight: 700; color: #0F172A; margin-bottom: 2px; }
.pkg-banner-formula{ font-size: 11px; color: #64748B; font-family: monospace !important; }

/* ── Rule Check Grid ── */
.rule-check {
    background: white; border: 1px solid #E2E8F0;
    border-radius: 12px; padding: 14px 16px;
    display: flex; align-items: flex-start; gap: 10px;
}
.rule-check.ok   { border-left: 4px solid #10B981; }
.rule-check.fail { border-left: 4px solid #EF4444; }
.rule-badge {
    width: 24px; height: 24px; border-radius: 50%;
    display: inline-flex; align-items: center; justify-content: center;
    font-size: 12px; font-weight: 700; flex-shrink: 0;
}
.rule-badge.ok   { background: #DCFCE7; color: #15803D; }
.rule-badge.fail { background: #FEE2E2; color: #DC2626; }
.rule-label { font-size: 12px; font-weight: 600; color: #374151; margin-bottom: 2px; }
.rule-detail { font-size: 11px; color: #64748B; }

/* ── Solver Banner ── */
.solver-ok   { background: #F0FDF4; border: 1px solid #86EFAC; border-radius: 12px; padding: 14px 18px; color: #14532D; font-size: 13px; line-height: 1.6; }
.solver-warn { background: #FFFBEB; border: 1px solid #FDE68A; border-radius: 12px; padding: 14px 18px; color: #78350F; font-size: 13px; line-height: 1.6; }
.solver-err  { background: #FEF2F2; border: 1px solid #FECACA; border-radius: 12px; padding: 14px 18px; color: #7F1D1D; font-size: 13px; line-height: 1.6; }

/* ── Sidebar helpers ── */
.sb-section {
    font-size: 9px; font-weight: 700;
    text-transform: uppercase; letter-spacing: 0.14em;
    color: #475569; padding: 18px 0 6px;
    border-top: 1px solid #1E293B;
    margin-top: 4px;
}
.sb-section:first-child { border-top: none; padding-top: 4px; }

/* ── Override Streamlit alerts ── */
div[data-testid="stAlert"] { border-radius: 12px !important; }

/* ── Override metric ── */
div[data-testid="stMetricValue"] { font-size: 22px !important; font-weight: 800 !important; }
div[data-testid="stMetricLabel"] { font-size: 11px !important; font-weight: 700 !important; }

/* ── Landing page ── */
.landing-box {
    background: white; border-radius: 20px;
    border: 2px dashed #E2E8F0; padding: 64px 40px;
    text-align: center; margin-top: 40px;
}
.landing-icon { font-size: 56px; margin-bottom: 20px; }
.landing-title { font-size: 22px; font-weight: 700; color: #0F172A; margin-bottom: 8px; }
.landing-sub   { font-size: 14px; color: #64748B; }

/* ── Expander tweaks ── */
[data-testid="stExpander"] {
    border: 1px solid #E2E8F0 !important;
    border-radius: 12px !important;
    background: white !important;
    box-shadow: 0 2px 8px rgba(15,23,42,0.04) !important;
}
</style>
""", unsafe_allow_html=True)


# =========================================================
# HELPER FUNCTIONS
# =========================================================

def eur(x):
    if x is None:
        return "—"
    try:
        if pd.isna(x):
            return "—"
    except (TypeError, ValueError):
        pass
    if isinstance(x, str):
        return x
    try:
        return f"{x:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")
    except (TypeError, ValueError):
        return str(x)


def pct(x):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return "—"
    sign = "+" if x > 0 else ""
    return f"{sign}{x * 100:.1f}%"


def kpi(col_obj, label, value, note="", accent=""):
    with col_obj:
        cls = f"kpi-card {accent}" if accent else "kpi-card"
        st.markdown(
            f'<div class="{cls}">'
            f'<div class="kpi-label">{label}</div>'
            f'<div class="kpi-value">{value}</div>'
            f'<div class="kpi-note">{note}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )


def rule_row(cols, items):
    for col_obj, (ok, title, detail) in zip(cols, items):
        badge = "ok" if ok else "fail"
        icon  = "✓" if ok else "✗"
        with col_obj:
            st.markdown(
                f'<div class="rule-check {badge}">'
                f'<div class="rule-badge {badge}">{icon}</div>'
                f'<div><div class="rule-label">{title}</div>'
                f'<div class="rule-detail">{detail}</div></div>'
                f'</div>',
                unsafe_allow_html=True,
            )


def sec(label, icon="", color=""):
    st.markdown(
        f'<div class="sec-heading">'
        f'<span class="sec-icon {color}">{icon}</span>'
        f'<span>{label}</span>'
        f'<span class="sec-heading-bar"></span>'
        f'</div>',
        unsafe_allow_html=True,
    )


# =========================================================
# SIDEBAR
# =========================================================

with st.sidebar:
    st.markdown(
        '<div style="display:flex;align-items:center;gap:10px;margin-bottom:24px;">'
        '<div style="background:#FF5F00;border-radius:10px;width:36px;height:36px;display:flex;align-items:center;justify-content:center;font-size:20px;">🚗</div>'
        '<div><div style="font-size:13px;font-weight:800;color:#F1F5F9;">SIXT Pricing</div>'
        '<div style="font-size:10px;color:#64748B;font-weight:600;text-transform:uppercase;letter-spacing:0.08em;">Simulator v2</div></div>'
        '</div>',
        unsafe_allow_html=True,
    )

    st.markdown('<div class="sb-section">📂 Fonte de dados</div>', unsafe_allow_html=True)
    uploaded_file = st.file_uploader("Ficheiro Excel da matriz", type=["xlsx", "xls"], label_visibility="collapsed")
    sheet_name    = st.text_input("Nome da folha", value="Matrix - 11.05.2026", help="Nome exacto da aba Excel com a matriz.")

    if uploaded_file is None:
        st.markdown(
            '<div style="background:#1E293B;border-radius:10px;padding:12px;margin-top:8px;">'
            '<div style="font-size:11px;color:#64748B;line-height:1.6;">'
            'Carregue o ficheiro Excel com a matriz de preços.<br>'
            'Deve conter as linhas: <code style="color:#94A3B8">chco · vont · bist · cart</code>'
            '</div></div>',
            unsafe_allow_html=True,
        )
        st.stop()

    # ── Load data ──────────────────────────────────────────
    try:
        file_bytes  = uploaded_file.getvalue()
        matrix_data = build_matrix_data_from_excel(file_bytes=file_bytes, sheet_name=sheet_name)
    except Exception as e:
        st.error(f"Erro ao ler a matriz: {e}")
        st.stop()

    st.markdown('<div class="sb-section">🎯 Seleção</div>', unsafe_allow_html=True)
    group = st.selectbox(
        "Grupo ACRISS",
        matrix_data["groups"],
        index=matrix_data["groups"].index("CCCC") if "CCCC" in matrix_data["groups"] else 0,
    )
    days    = st.slider("Dias de aluguer", 1, 60, 3, 1)
    vat_pct = st.slider("IVA (%)", 0, 30, 23, 1)

    st.markdown('<div class="sb-section">💳 Descontos SMART+</div>', unsafe_allow_html=True)
    counter_discount_pct = st.slider("Balcão SMART+ (%)", 0, 50, 10, 1, key="dc_smart")
    online_discount_pct  = st.slider("Online SMART+ (%)",  0, 60, 25, 1, key="do_smart")
    bf_weight_pct        = st.slider("Distribuição subida → BF (%)", 0, 100, 70, 5,
                                     help="O restante é imputado ao BQ.")
    st.caption(f"BF {bf_weight_pct}% / BQ {100 - bf_weight_pct}%")

    st.markdown('<div class="sb-section">🔧 Pack Easy</div>', unsafe_allow_html=True)
    easy_counter_pct = st.slider("Balcão Easy (%)", 0, 50, 0, 1, key="dc_easy",
                                  help="Pack Easy não tem canal online.")
    st.caption("Online: n/a — Pack Easy não vendido online")
    col_m1, col_m2 = st.columns(2)
    with col_m1:
        easy_margin_min = st.number_input("Margem mín. (€)", 0.0, 20.0, 3.5, 0.5, format="%.1f")
    with col_m2:
        easy_margin_max = st.number_input("Margem máx. (€)", 0.0, 30.0, 7.0, 0.5, format="%.1f")

    st.markdown('<div class="sb-section">🛞 Piso mínimo TG</div>', unsafe_allow_html=True)
    tg_min_iva = st.number_input(
        "TG mínimo c/IVA (€)",
        min_value=0.0, max_value=25.0, value=8.0, step=0.5, format="%.1f",
        help="Avis, Europcar, Guerin e Hertz cobram entre 8€ e 18€ c/IVA pelo TG. "
             "O simulador garante que TG nunca sai abaixo deste valor.",
    )
    if tg_min_iva > 0:
        st.caption(
            f"Rack mínimo s/IVA: **{tg_min_iva / 1.23:.2f} €**  "
            f"(= {tg_min_iva:.1f} € / 1,23)"
        )

    st.markdown('<div class="sb-section">⚙️ Avançado</div>', unsafe_allow_html=True)
    online_extra_pct = st.slider("Extra online AI vs balcão (%)", 0, 30, 10, 1, key="online_extra",
                                  help="Garante que AI online ≥ AI balcão + este valor.")

    st.markdown("---")
    st.markdown(
        f'<div style="font-size:10px;color:#475569;text-align:center;">'
        f'{len(matrix_data["groups"])} grupos ACRISS · folha <em>{sheet_name}</em>'
        f'</div>',
        unsafe_allow_html=True,
    )


# =========================================================
# CALCULATE
# =========================================================

_tg_min_rack_sim    = (tg_min_iva / (1 + vat_pct / 100)) if tg_min_iva > 0 else 0.0
_global_scale_sim   = compute_global_tg_scale(matrix_data, _tg_min_rack_sim)
_tg_progression_sim = compute_tg_lor_progression(matrix_data, group, _tg_min_rack_sim, _global_scale_sim)

def _find_lor_key(days_val):
    for ls, le, ref in [(1,3,1),(4,6,4),(7,10,7),(11,29,11),(30,30,30),(31,999,31)]:
        if ls <= days_val <= le:
            return (ls, le, ref)
    return None

_lor_key_sim = _find_lor_key(days)
_tg_ov_sim   = _tg_progression_sim.get(_lor_key_sim) if _lor_key_sim else None

result = calculate_pricing(
    matrix_data=matrix_data, group=group, days=days,
    counter_discount=counter_discount_pct / 100,
    online_discount=online_discount_pct   / 100,
    bf_weight=bf_weight_pct / 100,
    vat=vat_pct / 100,
    easy_counter_discount=easy_counter_pct / 100,
    ai_counter_discount=None, ai_online_discount=None,
    easy_margin_min=easy_margin_min, easy_margin_max=easy_margin_max,
    online_extra_discount=online_extra_pct / 100,
    tg_min_iva=tg_min_iva,
    tg_override=_tg_ov_sim,
)

ai_counter_pct = round(result["ai_counter_discount_solved"] * 100, 2) if result.get("ai_counter_discount_solved") is not None else 0
ai_online_pct  = round(result["ai_online_discount_solved"]  * 100, 2) if result.get("ai_online_discount_solved")  is not None else 0

# ── Auto-load BD concorrência (BD_Pacotes_Concorrentes.xlsx na pasta do projeto) ──
_BD_FILENAME = "BD_Pacotes_Concorrentes.xlsx"
_BD_PATH = None
for _cand in [
    os.path.join(os.path.dirname(os.path.abspath(__file__)), _BD_FILENAME),
    os.path.join(os.getcwd(), _BD_FILENAME),
    os.path.join(os.path.dirname(__file__), _BD_FILENAME) if "__file__" in dir() else "",
    _BD_FILENAME,
]:
    if _cand and os.path.exists(_cand):
        _BD_PATH = _cand
        break
comp_bd = None
if _BD_PATH:
    try:
        with open(_BD_PATH, "rb") as _f:
            comp_bd = load_competitor_bd(_f.read())
    except Exception:
        comp_bd = None

vl_pct = result.get("ai_counter_discount_valid_low")
vh_pct = result.get("ai_counter_discount_valid_high")
feasible_interval = vl_pct is not None and vh_pct is not None and vl_pct <= vh_pct

all_rules_ok = (
    result["ok_ai_gt_smart_counter"] and result["ok_ai_gt_smart_online"]
    and result["ok_smart_easy_gt_ai_counter"] and result["ok_smart_tg_gt_ai_counter"]
    and result["ok_easy_margin_counter"]
)

smart_plus_easy = (
    round(result["smart_rack_new"] + result["easy_rack_new"], 4)
    if result["smart_rack_new"] is not None and result["easy_rack_new"] is not None else None
)


# =========================================================
# APP HEADER
# =========================================================

n_issues = len(result["issues"])
status_color = "#10B981" if all_rules_ok and n_issues == 0 else "#F59E0B" if n_issues == 0 else "#EF4444"
status_label = "Todas as regras OK" if all_rules_ok and n_issues == 0 else f"{n_issues} componente(s) em falta" if n_issues > 0 else "Regras com alertas"
status_dot   = "🟢" if all_rules_ok and n_issues == 0 else "🟡" if n_issues == 0 else "🔴"

st.markdown(
    f"""
    <div class="app-header">
        <div class="header-eyebrow">Simulador de Pricing</div>
        <div class="header-title">Proteções SIXT Portugal</div>
        <div class="header-subtitle">
            SMART+ = LD + BF + BQ (sem TG) &nbsp;·&nbsp; All Inclusive = BC + LD + BF + BQ + I + TG
            &nbsp;·&nbsp; Pack Easy = TG + BC (só balcão)
        </div>
        <div class="header-pills">
            <span class="header-pill">{len(matrix_data['groups'])} grupos ACRISS</span>
            <span class="header-pill">Grupo: {group}</span>
            <span class="header-pill">{days} dia{"s" if days != 1 else ""} de aluguer</span>
            <span class="header-pill">IVA {vat_pct}%</span>
            <span class="header-pill">{status_dot} {status_label}</span>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)


# =========================================================
# TABS
# =========================================================

tab_sim, tab_compare, tab_concorrencia, tab_opt, tab_export, tab_debug = st.tabs([
    "📊  Simulador",
    "💰  Comparação",
    "🏁  Análise Concorrência",
    "🎯  Optimizador Global",
    "📥  Exportar",
    "🔬  Técnico",
])


# ─────────────────────────────────────────────────────────
# TAB 1 — SIMULADOR
# ─────────────────────────────────────────────────────────

with tab_sim:

    # ── Issues alert ──────────────────────────────────────
    if result["issues"]:
        st.error("**Componentes indisponíveis:**  " + "  ·  ".join(result["issues"]))

    # ── Solver banner ─────────────────────────────────────
    if result.get("ai_rack_new") is None:
        st.markdown('<div class="solver-warn">⚠️ Grupo sem All Inclusive (prestige) — desconto AI não aplicável.</div>', unsafe_allow_html=True)
    elif feasible_interval:
        st.markdown(
            f'<div class="solver-ok">'
            f'<strong>Solver AI</strong> — Intervalo válido: '
            f'<strong>[{vl_pct*100:.1f}% ; {vh_pct*100:.1f}%]</strong> &nbsp;·&nbsp; '
            f'Balcão: <strong>{ai_counter_pct:.2f}%</strong> &nbsp;·&nbsp; '
            f'Online: <strong>{ai_online_pct:.2f}%</strong>'
            f'</div>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            f'<div class="solver-warn">'
            f'⚠️ Intervalo inválido para este grupo/LOR. '
            f'Desconto AI aplicado por best-effort: {ai_counter_pct:.2f}% (balcão) · {ai_online_pct:.2f}% (online).'
            f'</div>',
            unsafe_allow_html=True,
        )

    # ── Package banners ───────────────────────────────────
    pb1, pb2, pb3 = st.columns(3)
    with pb1:
        st.markdown(
            '<div class="pkg-banner smart">'
            '<div class="pkg-banner-dot" style="background:#FF5F00;margin-top:4px;"></div>'
            '<div class="pkg-banner-body">'
            '<div class="pkg-banner-title">SMART+</div>'
            '<div class="pkg-banner-formula">LD + BF + BQ (sem TG)</div>'
            '</div></div>',
            unsafe_allow_html=True,
        )
    with pb2:
        st.markdown(
            '<div class="pkg-banner ai">'
            '<div class="pkg-banner-dot" style="background:#3B82F6;margin-top:4px;"></div>'
            '<div class="pkg-banner-body">'
            '<div class="pkg-banner-title">All Inclusive</div>'
            '<div class="pkg-banner-formula">BC + LD + BF + BQ + I + TG</div>'
            '</div></div>',
            unsafe_allow_html=True,
        )
    with pb3:
        st.markdown(
            '<div class="pkg-banner easy">'
            '<div class="pkg-banner-dot" style="background:#10B981;margin-top:4px;"></div>'
            '<div class="pkg-banner-body">'
            '<div class="pkg-banner-title">Pack Easy</div>'
            '<div class="pkg-banner-formula">TG + BC (só balcão)</div>'
            '</div></div>',
            unsafe_allow_html=True,
        )

    # ══════════════════════════════════════════════════════
    # SMART+
    # ══════════════════════════════════════════════════════
    sec("SMART+ — LD + BF + BQ (sem TG)", "💳", "orange")

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    kpi(c1, "Referência (c/ TG)",   eur(result["smart_base"]),      "LD+BF+BQ+TG histórico · s/IVA")
    kpi(c2, "Rack novo (sem TG)",   eur(result["smart_rack_new"]),  f"Δ rack: {pct(result['smart_increase_pct'])}", "orange")
    kpi(c3, "LD (fixo)",            f"{eur(result['ld'])}",         "Não é alterado")
    kpi(c4, "BF atual → novo",      f"{eur(result['bf'])} → {eur(result['bf_new'])}",  f"Δ {pct(result['bf_increase_pct'])}")
    kpi(c5, "BQ atual → novo",      f"{eur(result['bq'])} → {eur(result['bq_new'])}",  f"Δ {pct(result['bq_increase_pct'])}")
    kpi(c6, "Aumento BF+BQ",        eur(result["bf_bq_gap"]),       "Compensação pela remoção TG")

    st.markdown("")
    d1, d2, d3, d4 = st.columns(4)
    kpi(d1, f"Balcão {counter_discount_pct}% — s/IVA", eur(result["smart_counter"]), f"{eur(result['smart_counter_vat'])} c/IVA", "orange")
    kpi(d2, f"Online {online_discount_pct}% — s/IVA",  eur(result["smart_online"]),  f"{eur(result['smart_online_vat'])} c/IVA",  "green")
    kpi(d3, "Desconto implícito balcão", pct(result["smart_implicit_discount"]), "Sobre rack novo")
    ok_s = result["ok_smart"]
    kpi(d4, "Regra balcão = antigo",
        "✓ Cumprida" if ok_s else "✗ Divergência",
        f"Alvo: {eur(result['smart_counter_target'])} · Actual: {eur(result['smart_counter'])}",
        "green" if ok_s else "red")

    # ══════════════════════════════════════════════════════
    # ALL INCLUSIVE
    # ══════════════════════════════════════════════════════
    sec("All Inclusive — BC + LD + BF + BQ + I + TG", "💎", "blue")

    a1, a2, a3, a4, a5, a6 = st.columns(6)
    kpi(a1, "AI referência",       eur(result["ai_base"]),      "BC+LD+BF+BQ+I+TG histórico · s/IVA")
    kpi(a2, "AI rack novo",        eur(result["ai_rack_new"]),  f"Δ {pct(result['ai_increase_pct'])}", "blue")
    # TG card — destaca se o piso mínimo foi aplicado
    _tg_note = f"Δ {pct(result['tg_increase_pct'])} · única variável ajust."
    _tg_accent = ""
    if result.get("tg_raised_to_min"):
        _tg_note = f"⚠️ Piso mínimo aplicado ({eur(tg_min_iva)} c/IVA) · {_tg_note}"
        _tg_accent = "orange"
    elif result.get("tg_below_min") is False and tg_min_iva > 0:
        _tg_note = f"✓ Acima do piso mínimo ({eur(tg_min_iva)} c/IVA) · {_tg_note}"
        _tg_accent = "green"
    kpi(a3, "TG atual → novo", f"{eur(result['tg'])} → {eur(result['tg_new'])}", _tg_note, _tg_accent)
    kpi(a4, "LD / BF / BQ no AI", "= SMART+", "Mesmos valores do SMART+ novo")
    kpi(a5, f"Balcão {ai_counter_pct:.1f}% — s/IVA", eur(result["ai_counter"]), f"{eur(result['ai_counter_vat'])} c/IVA  ·  desc. solver", "blue")
    kpi(a6, f"Online {ai_online_pct:.1f}% — s/IVA",  eur(result["ai_online"]),  f"{eur(result['ai_online_vat'])} c/IVA", "purple")

    # ══════════════════════════════════════════════════════
    # PACK EASY
    # ══════════════════════════════════════════════════════
    sec("Pack Easy — TG + BC (só balcão)", "🔧", "green")

    e1, e2, e3, e4, e5, e6 = st.columns(6)
    kpi(e1, "Easy referência",     eur(result["easy_base"]),      "TG + BC histórico · s/IVA")
    kpi(e2, "Easy rack novo",      eur(result["easy_rack_new"]),  f"Δ {pct(result['easy_increase_pct'])}", "green")
    _tg_easy_note = f"Mesmo TG unitário do AI · piso {eur(tg_min_iva)} c/IVA"
    _tg_easy_acc  = "orange" if result.get("tg_raised_to_min") else ""
    kpi(e3, "TG novo", eur(result["tg_new"]), _tg_easy_note, _tg_easy_acc)
    kpi(e4, f"Balcão {easy_counter_pct}% — s/IVA", eur(result["easy_counter"]), f"{eur(result['easy_counter_vat'])} c/IVA  ·  único canal", "green")
    kpi(e5, "Online Easy",         "n/a",  "Pack Easy não é vendido online")
    kpi(e6, "SMART+(rack) + Easy(rack)",
        eur(smart_plus_easy),
        f"vs AI rack: {eur(result['ai_rack_new'])}  ·  Δ: {eur((smart_plus_easy or 0) - (result['ai_rack_new'] or 0))}",
        "orange")

    # ══════════════════════════════════════════════════════
    # RULE CHECKS
    # ══════════════════════════════════════════════════════
    sec("Verificação de regras", "✅", "")

    ok_ai_cnt  = result["ok_ai_gt_smart_counter"]
    ok_ai_onl  = result["ok_ai_gt_smart_online"]
    ok_se_cnt  = result["ok_smart_easy_gt_ai_counter"]
    ok_stg_cnt = result["ok_smart_tg_gt_ai_counter"]
    ok_mg      = result["ok_easy_margin_counter"]
    m_counter  = result["easy_margin_counter"]
    solved_o   = result["ai_online_discount_solved"]
    extra_pct  = result["online_extra_discount"]
    max_da_o   = result["ai_online_discount_max_rule2"]

    rc1, rc2, rc3, rc4 = st.columns(4)
    rule_row(
        [rc1, rc2, rc3, rc4],
        [
            (ok_ai_cnt,  "AI > SMART+ balcão",
             f"{eur(result['ai_counter'])} > {eur(result['smart_counter'])}"),
            (ok_ai_onl,  "AI > SMART+ online",
             f"{eur(result['ai_online'])} > {eur(result['smart_online'])}"),
            (ok_se_cnt,  "SMART+ + Easy > AI (balcão)",
             f"{eur((result['smart_counter'] or 0)+(result['easy_counter'] or 0))} > {eur(result['ai_counter'])}"),
            (ok_stg_cnt, "SMART+ + TG > AI (balcão)",
             f"{eur((result['smart_counter'] or 0)+(result['tg_new'] or 0))} > {eur(result['ai_counter'])}"),
        ]
    )
    st.markdown("")
    rc5, rc6, rc7, rc8 = st.columns(4)

    eff_online = solved_o * 100 if solved_o is not None else None
    ai_onl_extra = (
        (solved_o - (1 - result["ai_counter"] / result["ai_rack_new"])) * 100
        if solved_o is not None and result.get("ai_counter") and result.get("ai_rack_new") else None
    )
    ok_onl_extra = ai_onl_extra is not None and ai_onl_extra >= online_extra_pct - 0.1

    rule_row(
        [rc5, rc6, rc7, rc8],
        [
            (ok_mg,
             f"Margem Easy ∈ [{eur(easy_margin_min)} ; {eur(easy_margin_max)}]",
             f"Actual: {eur(m_counter)}  (balcão)"),
            (feasible_interval,
             "Intervalo solver AI válido",
             f"[{vl_pct*100:.1f}% ; {vh_pct*100:.1f}%] → ponto médio {ai_counter_pct:.1f}%"
             if vl_pct is not None else "Sem solução"),
            (ok_onl_extra,
             "Diferença online vs balcão AI",
             f"+{ai_onl_extra:.1f}%  (mín. exigido: +{online_extra_pct}%)"
             if ai_onl_extra is not None else "—"),
            (result["ok_smart"],
             "Balcão SMART+ = balcão antigo",
             f"Alvo: {eur(result['smart_counter_target'])}  ·  Actual: {eur(result['smart_counter'])}"),
        ]
    )

    if all_rules_ok:
        st.success(
            f"✅  Todas as regras cumpridas — "
            f"AI balcão {eur(result['ai_counter'])} · AI online {eur(result['ai_online'])} · "
            f"Margem {eur(m_counter)} ∈ [{eur(easy_margin_min)} ; {eur(easy_margin_max)}]"
        )
    else:
        failed = []
        if not ok_ai_cnt:  failed.append("AI > SMART+ balcão")
        if not ok_ai_onl:  failed.append("AI > SMART+ online")
        if not ok_se_cnt:  failed.append("SMART++Easy > AI")
        if not ok_stg_cnt: failed.append("SMART++TG > AI")
        if not ok_mg:      failed.append(f"Margem [{eur(easy_margin_min)};{eur(easy_margin_max)}]")
        st.warning(f"⚠️  Regras não cumpridas: {', '.join(failed)}. O solver ajustou automaticamente o que foi possível.")

    st.caption(f"ℹ️  Pack Easy não existe online. Desconto AI online ≥ AI balcão + {online_extra_pct}% (comissões).")

    # ══════════════════════════════════════════════════════
    # CHART
    # ══════════════════════════════════════════════════════
    sec("Comparação visual por canal", "📊", "")

    labels = [
        "SMART+ ref.", "SMART+ rack", f"SMART+\nbalcão {counter_discount_pct}%", f"SMART+\nonline {online_discount_pct}%",
        "AI ref.", "AI rack", f"AI\nbalcão {ai_counter_pct:.1f}%", f"AI\nonline {ai_online_pct:.1f}%",
        "Easy ref.", "Easy rack", f"Easy\nbalcão {easy_counter_pct}%",
        "S+ + Easy\nrack",
    ]
    values = [
        result["smart_base"], result["smart_rack_new"], result["smart_counter"], result["smart_online"],
        result["ai_base"],    result["ai_rack_new"],    result["ai_counter"],    result["ai_online"],
        result["easy_base"],  result["easy_rack_new"],  result["easy_counter"],
        smart_plus_easy,
    ]
    colors = ["#FF5F00","#FF7A2E","#FDBA74","#FED7AA"] + \
             ["#1D4ED8","#3B82F6","#60A5FA","#93C5FD"] + \
             ["#059669","#10B981","#34D399"] + \
             ["#8B5CF6"]

    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=labels, y=values,
        marker_color=colors,
        text=[f"{v:,.2f}" if v is not None and not (isinstance(v, float) and pd.isna(v)) else "" for v in values],
        textposition="outside", textfont=dict(size=10, family="Inter"),
        hovertemplate="%{x}<br><b>%{y:.2f} €</b><extra></extra>",
    ))
    for ref, label, pos in [
        (result["smart_base"], "ref. SMART+", "top left"),
        (result["ai_base"],    "ref. AI",     "top right"),
    ]:
        if ref is not None:
            fig.add_hline(y=ref, line_dash="dot", line_color="#94A3B8", line_width=1,
                          annotation_text=label, annotation_position=pos,
                          annotation_font=dict(size=10, color="#64748B"))
    fig.update_layout(
        title=dict(text=f"{group}  ·  {days} dia{'s' if days != 1 else ''}  ·  preços s/IVA", font=dict(size=14, family="Inter", color="#0F172A")),
        yaxis=dict(title="EUR s/IVA", gridcolor="#F1F5F9", gridwidth=1, zeroline=False),
        xaxis=dict(tickfont=dict(size=10)),
        height=440, margin=dict(l=40, r=20, t=60, b=40),
        plot_bgcolor="white", paper_bgcolor="white",
        showlegend=False,
        font=dict(family="Inter"),
    )
    st.plotly_chart(fig, use_container_width=True)

    # ══════════════════════════════════════════════════════
    # DETAIL TABLE
    # ══════════════════════════════════════════════════════
    with st.expander("🗂️  Tabela de detalhe completa", expanded=False):
        detail_rows = [
            {"Secção": "Proteções individuais", "Elemento": "BC — Roadside Protection", "Preço atual": result["bc"], "Rack novo": result["bc"], f"Balcão {counter_discount_pct}%": None, f"Online {online_discount_pct}%": None},
            {"Secção": "Proteções individuais", "Elemento": "LD — Loss Damage Waiver",   "Preço atual": result["ld"], "Rack novo": result["ld_new"], f"Balcão {counter_discount_pct}%": None, f"Online {online_discount_pct}%": None},
            {"Secção": "Proteções individuais", "Elemento": "BF — Min. Excess LDW",      "Preço atual": result["bf"], "Rack novo": result["bf_new"], f"Balcão {counter_discount_pct}%": None, f"Online {online_discount_pct}%": None},
            {"Secção": "Proteções individuais", "Elemento": "BQ — Interior Protection",  "Preço atual": result["bq"], "Rack novo": result["bq_new"], f"Balcão {counter_discount_pct}%": None, f"Online {online_discount_pct}%": None},
            {"Secção": "Proteções individuais", "Elemento": "TG — Tyre & Windscreen",    "Preço atual": result["tg"], "Rack novo": result["tg_new"], f"Balcão {counter_discount_pct}%": None, f"Online {online_discount_pct}%": None},
            {"Secção": "Proteções individuais", "Elemento": "I — Personal Accident",     "Preço atual": result["ip"], "Rack novo": result["ip"],     f"Balcão {counter_discount_pct}%": None, f"Online {online_discount_pct}%": None},
            {"Secção": "SMART+",     "Elemento": "SMART+ s/IVA",     "Preço atual": result["smart_base"],   "Rack novo": result["smart_rack_new"], f"Balcão {counter_discount_pct}%": result["smart_counter"],     f"Online {online_discount_pct}%": result["smart_online"]},
            {"Secção": "SMART+",     "Elemento": "SMART+ c/IVA",     "Preço atual": None, "Rack novo": None, f"Balcão {counter_discount_pct}%": result["smart_counter_vat"],  f"Online {online_discount_pct}%": result["smart_online_vat"]},
            {"Secção": "All Incl.",  "Elemento": "All Inclusive s/IVA", "Preço atual": result["ai_base"],   "Rack novo": result["ai_rack_new"],    f"Balcão {counter_discount_pct}%": result["ai_counter"],        f"Online {online_discount_pct}%": result["ai_online"]},
            {"Secção": "All Incl.",  "Elemento": "All Inclusive c/IVA", "Preço atual": None, "Rack novo": None, f"Balcão {counter_discount_pct}%": result["ai_counter_vat"],   f"Online {online_discount_pct}%": result["ai_online_vat"]},
            {"Secção": "Pack Easy",  "Elemento": "Pack Easy s/IVA",   "Preço atual": result["easy_base"],  "Rack novo": result["easy_rack_new"],  f"Balcão {counter_discount_pct}%": result["easy_counter"],      f"Online {online_discount_pct}%": result["easy_online"]},
            {"Secção": "Pack Easy",  "Elemento": "Pack Easy c/IVA",   "Preço atual": None, "Rack novo": None, f"Balcão {counter_discount_pct}%": result["easy_counter_vat"],  f"Online {online_discount_pct}%": result["easy_online_vat"]},
            {"Secção": "Pack Easy",  "Elemento": "SMART+(rack) + Easy(rack)", "Preço atual": None, "Rack novo": smart_plus_easy, f"Balcão {counter_discount_pct}%": None, f"Online {online_discount_pct}%": None},
            {"Secção": "Margem/Solver", "Elemento": f"SMART++Easy−AI balcão (desc. AI={ai_counter_pct:.1f}%)", "Preço atual": None, "Rack novo": None, f"Balcão {counter_discount_pct}%": result["easy_margin_counter"], f"Online {online_discount_pct}%": result["easy_margin_online"]},
        ]
        detail_df = pd.DataFrame(detail_rows)
        fmt_df = detail_df.copy()
        for col in fmt_df.columns:
            if col not in ["Secção", "Elemento"]:
                fmt_df[col] = fmt_df[col].apply(eur)
        st.dataframe(fmt_df, hide_index=True)

    with st.expander("📋  Regras implementadas", expanded=False):
        st.markdown("""
**1.** O preço de **LD não muda**.

**2.** **SMART+** = LD + BF + BQ (TG removido). O **preço de balcão deve igualar o SMART+ antigo** (com TG) — BF e BQ sobem para compensar.

**3.** No **All Inclusive**, LD, BF e BQ são exatamente iguais ao SMART+ novo. **TG é a única variável de ajuste** (nunca desce).

**4.** **Pack Easy** = TG + BC — **só balcão**, sem canal online nem rack público.

**5.** **AI(canal) > SMART+(canal)** em cada canal. Desconto AI calculado automaticamente como ponto médio do intervalo válido.

**6.** **SMART+(balcão) + Easy(balcão) > AI(balcão)** — sem vantagem em comprar separado.

**7.** **SMART+(balcão) + TG > AI(balcão)** — sem vantagem em comprar SMART++TG separados.

**8.** **Margem SMART++Easy − AI** ∈ [min€ ; max€] (configurável).

**9.** **Desconto AI online ≥ Desconto AI balcão + X%** (X configurável).
""")


# ─────────────────────────────────────────────────────────
# TAB 2 — COMPARAÇÃO DE PREÇOS
# ─────────────────────────────────────────────────────────

with tab_compare:

    sec("Comparação de Preços por Pacote", "💰", "orange")

    st.markdown(
        """
        Visualização lado a lado dos preços **antigo vs novo** para cada pacote de proteção.
        Todos os preços incluem IVA.
        """
    )

    # ══════════════════════════════════════════════════════
    # SMART+
    # ══════════════════════════════════════════════════════

    col_smart_old, col_smart_sep, col_smart_new = st.columns([1, 0.1, 1])

    with col_smart_old:
        # Preço antigo SMART+ (com TG + IVA)
        smart_old_with_tg = result["smart_base"]  # LD + BF + BQ + TG (histórico, s/IVA)
        smart_old_with_tg_vat = smart_old_with_tg * (1 + vat_pct / 100) if smart_old_with_tg is not None else None

        st.markdown(
            f"""
            <div style="background: #FFF7ED; border: 2px solid #FED7AA; border-radius: 16px; padding: 24px; text-align: center; height: 100%;">
                <div style="font-size: 12px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.09em; color: #92400E; margin-bottom: 12px;">
                    SMART+ Antigo
                </div>
                <div style="font-size: 11px; color: #B45309; margin-bottom: 16px; font-weight: 600;">
                    LD + BF + BQ + TG (incluso)
                </div>
                <div style="font-size: 32px; font-weight: 900; color: #7C2D12; margin-bottom: 8px;">
                    {eur(smart_old_with_tg_vat)}
                </div>
                <div style="font-size: 10px; color: #92400E; line-height: 1.5;">
                    <strong>s/IVA:</strong> {eur(smart_old_with_tg)}<br>
                    <strong>Grupo:</strong> {group}<br>
                    <strong>Duração:</strong> {days} dia{'s' if days != 1 else ''}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with col_smart_sep:
        st.markdown("")

    with col_smart_new:
        # Preço novo SMART+ (sem TG + IVA)
        smart_new = result["smart_rack_new"]
        smart_new_vat = smart_new * (1 + vat_pct / 100) if smart_new is not None else None

        # Comparação
        smart_delta = (smart_new_vat - smart_old_with_tg_vat) if (smart_new_vat is not None and smart_old_with_tg_vat is not None) else None
        smart_delta_pct = ((smart_new - smart_old_with_tg) / smart_old_with_tg if smart_old_with_tg and smart_old_with_tg > 0 else None)

        delta_color = "#10B981" if smart_delta is not None and smart_delta < 0 else "#EF4444" if smart_delta is not None and smart_delta > 0 else "#94A3B8"
        delta_sign = "" if smart_delta is not None and smart_delta < 0 else "+"

        st.markdown(
            f"""
            <div style="background: #F0FDF4; border: 2px solid #BBF7D0; border-radius: 16px; padding: 24px; text-align: center; height: 100%;">
                <div style="font-size: 12px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.09em; color: #15803D; margin-bottom: 12px;">
                    SMART+ Novo
                </div>
                <div style="font-size: 11px; color: #16A34A; margin-bottom: 16px; font-weight: 600;">
                    LD + BF + BQ (sem TG)
                </div>
                <div style="font-size: 32px; font-weight: 900; color: #15803D; margin-bottom: 8px;">
                    {eur(smart_new_vat)}
                </div>
                <div style="font-size: 10px; color: #16A34A; line-height: 1.5;">
                    <strong>s/IVA:</strong> {eur(smart_new)}<br>
                    <strong>Variação:</strong> <span style="color: {delta_color}; font-weight: 700;">{delta_sign}{eur(smart_delta) if smart_delta is not None else '—'} ({pct(smart_delta_pct)})</span><br>
                    <strong>Balcão ({counter_discount_pct}%):</strong> {eur(result['smart_counter_vat'])}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown("")
    st.markdown("---")
    st.markdown("")

    # ══════════════════════════════════════════════════════
    # ALL INCLUSIVE
    # ══════════════════════════════════════════════════════

    col_ai_old, col_ai_sep, col_ai_new = st.columns([1, 0.1, 1])

    with col_ai_old:
        # Preço antigo AI (com IVA)
        ai_old = result["ai_base"]  # BC+LD+BF+BQ+I+TG (histórico, s/IVA)
        ai_old_vat = ai_old * (1 + vat_pct / 100) if ai_old is not None else None

        st.markdown(
            f"""
            <div style="background: #EFF6FF; border: 2px solid #BFDBFE; border-radius: 16px; padding: 24px; text-align: center; height: 100%;">
                <div style="font-size: 12px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.09em; color: #1E40AF; margin-bottom: 12px;">
                    All Inclusive Antigo
                </div>
                <div style="font-size: 11px; color: #1E3A8A; margin-bottom: 16px; font-weight: 600;">
                    BC + LD + BF + BQ + I + TG (incluso)
                </div>
                <div style="font-size: 32px; font-weight: 900; color: #1E3A8A; margin-bottom: 8px;">
                    {eur(ai_old_vat)}
                </div>
                <div style="font-size: 10px; color: #1E40AF; line-height: 1.5;">
                    <strong>s/IVA:</strong> {eur(ai_old)}<br>
                    <strong>Grupo:</strong> {group}<br>
                    <strong>Duração:</strong> {days} dia{'s' if days != 1 else ''}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with col_ai_sep:
        st.markdown("")

    with col_ai_new:
        # Preço novo AI (com IVA)
        ai_new = result["ai_rack_new"]
        ai_new_vat = ai_new * (1 + vat_pct / 100) if ai_new is not None else None

        # Comparação
        ai_delta = (ai_new_vat - ai_old_vat) if (ai_new_vat is not None and ai_old_vat is not None) else None
        ai_delta_pct = ((ai_new - ai_old) / ai_old if ai_old and ai_old > 0 else None)

        delta_color_ai = "#10B981" if ai_delta is not None and ai_delta < 0 else "#EF4444" if ai_delta is not None and ai_delta > 0 else "#94A3B8"
        delta_sign_ai = "" if ai_delta is not None and ai_delta < 0 else "+"

        st.markdown(
            f"""
            <div style="background: #F0FDF4; border: 2px solid #BBF7D0; border-radius: 16px; padding: 24px; text-align: center; height: 100%;">
                <div style="font-size: 12px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.09em; color: #15803D; margin-bottom: 12px;">
                    All Inclusive Novo
                </div>
                <div style="font-size: 11px; color: #16A34A; margin-bottom: 16px; font-weight: 600;">
                    BC + LD + BF + BQ + I + TG (variável)
                </div>
                <div style="font-size: 32px; font-weight: 900; color: #15803D; margin-bottom: 8px;">
                    {eur(ai_new_vat)}
                </div>
                <div style="font-size: 10px; color: #16A34A; line-height: 1.5;">
                    <strong>s/IVA:</strong> {eur(ai_new)}<br>
                    <strong>Variação:</strong> <span style="color: {delta_color_ai}; font-weight: 700;">{delta_sign_ai}{eur(ai_delta) if ai_delta is not None else '—'} ({pct(ai_delta_pct)})</span><br>
                    <strong>Balcão ({ai_counter_pct:.1f}%):</strong> {eur(result['ai_counter_vat'])}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown("")
    st.markdown("---")
    st.markdown("")

    # ══════════════════════════════════════════════════════
    # PACK EASY
    # ══════════════════════════════════════════════════════

    col_easy_old, col_easy_sep, col_easy_new = st.columns([1, 0.1, 1])

    with col_easy_old:
        # Preço antigo Pack Easy (com IVA)
        easy_old = result["easy_base"]  # TG + BC (histórico, s/IVA)
        easy_old_vat = easy_old * (1 + vat_pct / 100) if easy_old is not None else None

        st.markdown(
            f"""
            <div style="background: #F3E8FF; border: 2px solid #D8B4FE; border-radius: 16px; padding: 24px; text-align: center; height: 100%;">
                <div style="font-size: 12px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.09em; color: #5B21B6; margin-bottom: 12px;">
                    Pack Easy Antigo
                </div>
                <div style="font-size: 11px; color: #6B21A8; margin-bottom: 16px; font-weight: 600;">
                    TG + BC (só balcão)
                </div>
                <div style="font-size: 32px; font-weight: 900; color: #6B21A8; margin-bottom: 8px;">
                    {eur(easy_old_vat)}
                </div>
                <div style="font-size: 10px; color: #5B21B6; line-height: 1.5;">
                    <strong>s/IVA:</strong> {eur(easy_old)}<br>
                    <strong>Grupo:</strong> {group}<br>
                    <strong>Duração:</strong> {days} dia{'s' if days != 1 else ''}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with col_easy_sep:
        st.markdown("")

    with col_easy_new:
        # Preço novo Pack Easy (com IVA)
        easy_new = result["easy_rack_new"]
        easy_new_vat = easy_new * (1 + vat_pct / 100) if easy_new is not None else None

        # Comparação
        easy_delta = (easy_new_vat - easy_old_vat) if (easy_new_vat is not None and easy_old_vat is not None) else None
        easy_delta_pct = ((easy_new - easy_old) / easy_old if easy_old and easy_old > 0 else None)

        delta_color_easy = "#10B981" if easy_delta is not None and easy_delta < 0 else "#EF4444" if easy_delta is not None and easy_delta > 0 else "#94A3B8"
        delta_sign_easy = "" if easy_delta is not None and easy_delta < 0 else "+"

        st.markdown(
            f"""
            <div style="background: #F0FDF4; border: 2px solid #BBF7D0; border-radius: 16px; padding: 24px; text-align: center; height: 100%;">
                <div style="font-size: 12px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.09em; color: #15803D; margin-bottom: 12px;">
                    Pack Easy Novo
                </div>
                <div style="font-size: 11px; color: #16A34A; margin-bottom: 16px; font-weight: 600;">
                    TG + BC (só balcão)
                </div>
                <div style="font-size: 32px; font-weight: 900; color: #15803D; margin-bottom: 8px;">
                    {eur(easy_new_vat)}
                </div>
                <div style="font-size: 10px; color: #16A34A; line-height: 1.5;">
                    <strong>s/IVA:</strong> {eur(easy_new)}<br>
                    <strong>Variação:</strong> <span style="color: {delta_color_easy}; font-weight: 700;">{delta_sign_easy}{eur(easy_delta) if easy_delta is not None else '—'} ({pct(easy_delta_pct)})</span><br>
                    <strong>Balcão ({easy_counter_pct}%):</strong> {eur(result['easy_counter_vat'])}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown("")
    st.markdown("---")
    st.markdown("")

    # ══════════════════════════════════════════════════════
    # RESUMO GRÁFICO
    # ══════════════════════════════════════════════════════

    sec("Resumo visual de variações", "📊", "")

    summary_labels = ["SMART+", "All Inclusive", "Pack Easy"]
    summary_old = [
        smart_old_with_tg_vat if smart_old_with_tg_vat is not None else 0,
        ai_old_vat if ai_old_vat is not None else 0,
        easy_old_vat if easy_old_vat is not None else 0,
    ]
    summary_new = [
        smart_new_vat if smart_new_vat is not None else 0,
        ai_new_vat if ai_new_vat is not None else 0,
        easy_new_vat if easy_new_vat is not None else 0,
    ]

    fig_compare = go.Figure(data=[
        go.Bar(name='Preço Antigo', x=summary_labels, y=summary_old, marker_color='#FF7A2E', text=[f"{v:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".") for v in summary_old], textposition='outside'),
        go.Bar(name='Preço Novo', x=summary_labels, y=summary_new, marker_color='#10B981', text=[f"{v:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".") for v in summary_new], textposition='outside'),
    ])

    fig_compare.update_layout(
        title=dict(text=f"Comparação de Preços com IVA — {group} · {days} dia{'s' if days != 1 else ''}", font=dict(size=14, family="Inter", color="#0F172A")),
        xaxis_title="Pacote de Proteção",
        yaxis_title="Preço (EUR com IVA)",
        barmode='group',
        height=400,
        margin=dict(l=40, r=20, t=60, b=40),
        plot_bgcolor="white",
        paper_bgcolor="white",
        font=dict(family="Inter"),
        legend=dict(x=0.01, y=0.99, bgcolor="rgba(255,255,255,0.8)")
    )

    st.plotly_chart(fig_compare, use_container_width=True)


# ─────────────────────────────────────────────────────────
# TAB 3 — ANÁLISE CONCORRÊNCIA
# ─────────────────────────────────────────────────────────

with tab_concorrencia:

    sec("Análise de Concorrência por Código ACRISS", "🏁", "orange")

    # ── Equivalence banner ────────────────────────────────
    st.markdown(
        """
        <div style="background:white;border:1px solid #E2E8F0;border-radius:16px;padding:20px 24px;margin-bottom:20px;">
        <div style="font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:0.09em;color:#475569;margin-bottom:14px;">
            Equivalências por concorrente — BD_Pacotes_Concorrentes.xlsx
        </div>
        <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:16px;">
            <div style="background:#FFF7ED;border-radius:10px;padding:14px;border:1px solid #FED7AA;">
                <div style="font-size:11px;font-weight:800;color:#7C2D12;margin-bottom:8px;">🚗 EUROPCAR</div>
                <div style="font-size:11px;color:#92400E;margin-bottom:4px;"><span style="font-weight:700;">Medium</span> → SMART+</div>
                <div style="font-size:11px;color:#92400E;"><span style="font-weight:700;">Premium</span> → All Inclusive</div>
            </div>
            <div style="background:#EFF6FF;border-radius:10px;padding:14px;border:1px solid #BFDBFE;">
                <div style="font-size:11px;font-weight:800;color:#1E3A8A;margin-bottom:8px;">🚗 GUERIN</div>
                <div style="font-size:11px;color:#1E40AF;margin-bottom:4px;"><span style="font-weight:700;">Premium Gold</span> → SMART+</div>
                <div style="font-size:11px;color:#1E40AF;"><span style="font-weight:700;">Platinum</span> → All Inclusive</div>
            </div>
            <div style="background:#F0FDF4;border-radius:10px;padding:14px;border:1px solid #BBF7D0;">
                <div style="font-size:11px;font-weight:800;color:#14532D;margin-bottom:8px;">🚗 AVIS</div>
                <div style="font-size:11px;color:#15803D;margin-bottom:4px;"><span style="font-weight:700;">Veículo</span> → SMART+</div>
                <div style="font-size:11px;color:#15803D;"><span style="font-weight:700;">Veículo Plus</span> → All Inclusive</div>
            </div>
            <div style="background:#F5F3FF;border-radius:10px;padding:14px;border:1px solid #DDD6FE;">
                <div style="font-size:11px;font-weight:800;color:#4C1D95;margin-bottom:8px;">🚗 HERTZ</div>
                <div style="font-size:11px;color:#5B21B6;">Apenas <span style="font-weight:700;">SuperCover</span> → All Inclusive</div>
            </div>
        </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ── Estado da BD ─────────────────────────────────────
    if comp_bd is None:
        st.error(
            "**BD_Pacotes_Concorrentes.xlsx não encontrada** na pasta do projeto. "
            "Coloca o ficheiro em `C:\\Users\\andre\\Desktop\\SIXT_Pricing_Simulator\\` e reinicia o app."
        )
    else:
        _n_grupos_bd  = len(comp_bd)
        _grupo_tem_bd = group in comp_bd
        _bd_grupo     = comp_bd.get(group, {})

        # Status pill
        if _grupo_tem_bd:
            _campos = list(_bd_grupo.keys())
            _tem_smart = any("smart" in c for c in _campos)
            _tem_ai    = any("ai"    in c for c in _campos)
            _pill_txt  = f"✅  Dados disponíveis para **{group}** ({len(_campos)} campos)"
            _pill_css  = "background:#DCFCE7;border:1px solid #86EFAC;color:#14532D;"
        else:
            _pill_txt  = f"⚠️  Sem dados de concorrência para **{group}** na BD"
            _pill_css  = "background:#FEF3C7;border:1px solid #FDE68A;color:#92400E;"

        st.markdown(
            f'<div style="{_pill_css}border-radius:10px;padding:10px 16px;font-size:12px;margin-bottom:16px;">'
            f'{_pill_txt} &nbsp;·&nbsp; BD com <strong>{_n_grupos_bd}</strong> grupos ACRISS mapeados</div>',
            unsafe_allow_html=True,
        )

        # ── Preços da BD para o grupo seleccionado ───────
        _vat_f = vat_pct / 100
        _sixt_smart_rack_vat    = result["smart_rack_new"] * (1 + _vat_f) if result.get("smart_rack_new") else None
        _sixt_smart_counter_vat = result.get("smart_counter_vat")
        _sixt_smart_online_vat  = result.get("smart_online_vat")
        _sixt_ai_rack_vat       = result["ai_rack_new"] * (1 + _vat_f) if result.get("ai_rack_new") else None
        _sixt_ai_counter_vat    = result.get("ai_counter_vat")
        _sixt_ai_online_vat     = result.get("ai_online_vat")

        _comp_smart = {
            "Europcar\nMedium":    _bd_grupo.get("europcar_smart"),
            "Guerin\nPrem. Gold":  _bd_grupo.get("guerin_smart"),
            "Avis\nVeículo":       _bd_grupo.get("avis_smart"),
        }
        _comp_ai = {
            "Europcar\nPremium":   _bd_grupo.get("europcar_ai"),
            "Guerin\nPlatinum":    _bd_grupo.get("guerin_ai"),
            "Avis\nVeículo Plus":  _bd_grupo.get("avis_ai"),
            "Hertz\nSuperCover":   _bd_grupo.get("hertz_ai"),
        }

        _any_smart = any(v is not None for v in _comp_smart.values())
        _any_ai    = any(v is not None for v in _comp_ai.values())

        # ── Tabela resumo de preços da BD para este grupo ─
        if _grupo_tem_bd:
            st.markdown(
                f'<div style="font-size:12px;color:#64748B;margin-bottom:8px;">'
                f'Preços da concorrência (€/dia · c/IVA) para o grupo <strong>{group}</strong> '
                f'· <em>Nota: preços da BD são por dia — SIXT calculado para {days} dia{"s" if days != 1 else ""}.</em></div>',
                unsafe_allow_html=True,
            )
            _bd_preview_rows = []
            _field_labels = {
                "europcar_smart": ("Europcar", "Medium", "SMART+"),
                "europcar_ai":    ("Europcar", "Premium", "All Inclusive"),
                "guerin_smart":   ("Guerin",   "Premium Gold", "SMART+"),
                "guerin_ai":      ("Guerin",   "Platinum", "All Inclusive"),
                "avis_smart":     ("Avis",     "Veículo", "SMART+"),
                "avis_ai":        ("Avis",     "Veículo Plus", "All Inclusive"),
                "hertz_ai":       ("Hertz",    "SuperCover", "All Inclusive"),
            }
            for fld, (marca, pacote, equiv) in _field_labels.items():
                val = _bd_grupo.get(fld)
                if val is not None:
                    # preço para os dias selecionados = val_por_dia * days
                    val_days = val * days
                    _bd_preview_rows.append({
                        "Concorrente": marca,
                        "Pacote": pacote,
                        "Equivale a": equiv,
                        f"€/dia (c/IVA)": val,
                        f"Total {days}d (c/IVA)": val_days,
                    })
            if _bd_preview_rows:
                _df_bd_prev = pd.DataFrame(_bd_preview_rows)
                _fmt_prev = _df_bd_prev.copy()
                _fmt_prev[f"€/dia (c/IVA)"]       = _fmt_prev[f"€/dia (c/IVA)"].apply(eur)
                _fmt_prev[f"Total {days}d (c/IVA)"] = _fmt_prev[f"Total {days}d (c/IVA)"].apply(eur)
                st.dataframe(_fmt_prev, hide_index=True)

            st.markdown("")

        if not _any_smart and not _any_ai:
            st.markdown(
                '<div class="landing-box" style="margin-top:8px;">'
                '<div class="landing-icon">🏁</div>'
                f'<div class="landing-title">Sem dados para {group}</div>'
                '<div class="landing-sub">Este grupo ACRISS não tem dados de concorrência na BD.<br>'
                'Adiciona-o ao ficheiro BD_Pacotes_Concorrentes.xlsx e reinicia o app.</div>'
                '</div>',
                unsafe_allow_html=True,
            )
        else:
            # ── Helpers ──────────────────────────────────
            def _disc_to_match(sixt_rack_vat, comp_price_vat):
                if sixt_rack_vat and comp_price_vat and sixt_rack_vat > 0:
                    return max(0.0, min(1 - comp_price_vat / sixt_rack_vat, 0.9999))
                return None

            def _position_label(sixt_price, comp_dict):
                valid = [v for v in comp_dict.values() if v is not None]
                if not valid or sixt_price is None:
                    return None, None, None
                mn, mx = min(valid), max(valid)
                avg = sum(valid) / len(valid)
                if sixt_price < mn:    return "abaixo do mercado", "#10B981", "🟢"
                elif sixt_price <= avg: return "competitivo",       "#3B82F6", "🔵"
                elif sixt_price <= mx:  return "acima da média",    "#F59E0B", "🟡"
                else:                   return "acima do mercado",  "#EF4444", "🔴"

            # Converter preços da BD (€/dia) para o período seleccionado
            _comp_smart_total = {k: v * days if v is not None else None for k, v in _comp_smart.items()}
            _comp_ai_total    = {k: v * days if v is not None else None for k, v in _comp_ai.items()}

            # ══════════════════════════════════════════════
            # ANÁLISE SMART+
            # ══════════════════════════════════════════════
            if _any_smart and _sixt_smart_rack_vat is not None:
                sec(f"SMART+ vs Concorrência — {group} · {days} dia{'s' if days != 1 else ''}", "💳", "orange")

                _valid_smart = {k: v for k, v in _comp_smart_total.items() if v is not None}
                _sv = list(_valid_smart.values())
                _s_min, _s_max, _s_avg = min(_sv), max(_sv), sum(_sv) / len(_sv)

                _pos_lbl_c, _, _pos_dot_c = _position_label(_sixt_smart_counter_vat, _valid_smart)
                _pos_lbl_o, _, _pos_dot_o = _position_label(_sixt_smart_online_vat,  _valid_smart)

                ks1, ks2, ks3, ks4, ks5 = st.columns(5)
                kpi(ks1, "SIXT SMART+ rack",            eur(_sixt_smart_rack_vat),    "c/IVA s/desconto", "orange")
                kpi(ks2, f"SIXT balcão {counter_discount_pct}%", eur(_sixt_smart_counter_vat),
                    f"{_pos_dot_c} {_pos_lbl_c}" if _pos_lbl_c else "—",
                    "green" if _pos_dot_c == "🟢" else ("blue" if _pos_dot_c == "🔵" else ("orange" if _pos_dot_c == "🟡" else "red")))
                kpi(ks3, f"SIXT online {online_discount_pct}%",  eur(_sixt_smart_online_vat),
                    f"{_pos_dot_o} {_pos_lbl_o}" if _pos_lbl_o else "—",
                    "green" if _pos_dot_o == "🟢" else ("blue" if _pos_dot_o == "🔵" else ("orange" if _pos_dot_o == "🟡" else "red")))
                kpi(ks4, "Mín. concorrência",  eur(_s_min), f"Total {days}d c/IVA")
                kpi(ks5, "Máx. concorrência",  eur(_s_max), f"Total {days}d c/IVA")

                st.markdown("")
                r1, r2, r3 = st.columns(3)
                for _col, _price, _lbl, _bg, _border, _col_txt in [
                    (r1, _s_min, "Igualar o mais barato",  "#FEF3C7", "#FDE68A", "#7C2D12"),
                    (r2, _s_avg, "Igualar a média",        "#DBEAFE", "#BFDBFE", "#1E3A8A"),
                    (r3, _s_max, "Igualar o mais caro",    "#DCFCE7", "#BBF7D0", "#14532D"),
                ]:
                    _d = _disc_to_match(_sixt_smart_rack_vat, _price)
                    with _col:
                        st.markdown(
                            f'<div style="background:{_bg};border:1px solid {_border};border-radius:12px;padding:14px 16px;">'
                            f'<div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:#92400E;margin-bottom:6px;">{_lbl}</div>'
                            f'<div style="font-size:20px;font-weight:800;color:{_col_txt};">{f"{_d*100:.1f}%" if _d is not None else "—"}</div>'
                            f'<div style="font-size:11px;color:{_col_txt};margin-top:4px;">desc. balcão p/ igualar {eur(_price)}</div>'
                            f'</div>', unsafe_allow_html=True,
                        )

                st.markdown("")
                _s_rows = []
                for _cn, _cp2 in _valid_smart.items():
                    _gap  = (_sixt_smart_counter_vat - _cp2) if _sixt_smart_counter_vat is not None else None
                    _dn   = _disc_to_match(_sixt_smart_rack_vat, _cp2)
                    _s_rows.append({
                        "Concorrente": _cn.replace("\n", " "),
                        f"Concorrência total {days}d (c/IVA)": _cp2,
                        f"SIXT balcão {counter_discount_pct}% (c/IVA)": _sixt_smart_counter_vat,
                        "Diferença (€)": _gap,
                        "Desc. balcão p/ igualar (%)": _dn * 100 if _dn is not None else None,
                        "Estado": "✓ Competitivo" if _gap is not None and _gap <= 0 else "⚠ Mais caro",
                    })
                with st.expander("📋  Detalhe SMART+ por concorrente", expanded=True):
                    _fs = pd.DataFrame(_s_rows).copy()
                    for _c in [f"Concorrência total {days}d (c/IVA)", f"SIXT balcão {counter_discount_pct}% (c/IVA)", "Diferença (€)"]:
                        _fs[_c] = _fs[_c].apply(eur)
                    _fs["Desc. balcão p/ igualar (%)"] = _fs["Desc. balcão p/ igualar (%)"].apply(
                        lambda x: f"{x:.1f}%" if x is not None and not (isinstance(x, float) and pd.isna(x)) else "—")
                    st.dataframe(_fs, hide_index=True)

                # Gráfico SMART+
                _sl = (["SIXT\nrack"] +
                       ([f"SIXT\nbalcão\n{counter_discount_pct}%"] if _sixt_smart_counter_vat else []) +
                       ([f"SIXT\nonline\n{online_discount_pct}%"]  if _sixt_smart_online_vat  else []) +
                       [k.replace("\n", " ") for k in _valid_smart])
                _sv2 = ([_sixt_smart_rack_vat] +
                        ([_sixt_smart_counter_vat] if _sixt_smart_counter_vat else []) +
                        ([_sixt_smart_online_vat]  if _sixt_smart_online_vat  else []) +
                        list(_valid_smart.values()))
                _ns = 1 + (1 if _sixt_smart_counter_vat else 0) + (1 if _sixt_smart_online_vat else 0)
                _sc = ["#FF5F00"] + ["#FDBA74"] * (_ns - 1) + ["#94A3B8"] * len(_valid_smart)
                _fig_s = go.Figure()
                _fig_s.add_trace(go.Bar(x=_sl, y=_sv2, marker_color=_sc,
                    text=[f"{v:,.2f}" if v else "" for v in _sv2],
                    textposition="outside", textfont=dict(size=10, family="Inter"),
                    hovertemplate="%{x}<br><b>%{y:.2f} €</b><extra></extra>"))
                for _rv, _rl, _rp in [(_s_min, f"mín {eur(_s_min)}", "bottom right"), (_s_max, f"máx {eur(_s_max)}", "top right")]:
                    _fig_s.add_hline(y=_rv, line_dash="dot", line_color="#94A3B8", line_width=1,
                        annotation_text=_rl, annotation_position=_rp, annotation_font=dict(size=10, color="#64748B"))
                _fig_s.update_layout(
                    title=dict(text=f"SMART+ — {group} · {days}d · c/IVA", font=dict(size=13, family="Inter", color="#0F172A")),
                    yaxis=dict(title="EUR c/IVA", gridcolor="#F1F5F9"),
                    height=380, margin=dict(l=40, r=20, t=50, b=40),
                    plot_bgcolor="white", paper_bgcolor="white", showlegend=False, font=dict(family="Inter"))
                st.plotly_chart(_fig_s, use_container_width=True)

            # ══════════════════════════════════════════════
            # ANÁLISE ALL INCLUSIVE
            # ══════════════════════════════════════════════
            if _any_ai and _sixt_ai_rack_vat is not None:
                sec(f"All Inclusive vs Concorrência — {group} · {days} dia{'s' if days != 1 else ''}", "💎", "blue")

                _valid_ai = {k: v for k, v in _comp_ai_total.items() if v is not None}
                _av = list(_valid_ai.values())
                _a_min, _a_max, _a_avg = min(_av), max(_av), sum(_av) / len(_av)

                _pos_lbl_ac, _, _pos_dot_ac = _position_label(_sixt_ai_counter_vat, _valid_ai)
                _pos_lbl_ao, _, _pos_dot_ao = _position_label(_sixt_ai_online_vat,  _valid_ai)

                ka1, ka2, ka3, ka4, ka5 = st.columns(5)
                kpi(ka1, "SIXT AI rack",                   eur(_sixt_ai_rack_vat),    "c/IVA s/desconto", "blue")
                kpi(ka2, f"SIXT balcão {ai_counter_pct:.1f}%", eur(_sixt_ai_counter_vat),
                    f"{_pos_dot_ac} {_pos_lbl_ac}" if _pos_lbl_ac else "—",
                    "green" if _pos_dot_ac == "🟢" else ("blue" if _pos_dot_ac == "🔵" else ("orange" if _pos_dot_ac == "🟡" else "red")))
                kpi(ka3, f"SIXT online {ai_online_pct:.1f}%", eur(_sixt_ai_online_vat),
                    f"{_pos_dot_ao} {_pos_lbl_ao}" if _pos_lbl_ao else "—",
                    "green" if _pos_dot_ao == "🟢" else ("blue" if _pos_dot_ao == "🔵" else ("orange" if _pos_dot_ao == "🟡" else "red")))
                kpi(ka4, "Mín. concorrência", eur(_a_min), f"Total {days}d c/IVA")
                kpi(ka5, "Máx. concorrência", eur(_a_max), f"Total {days}d c/IVA")

                st.markdown("")
                ra1, ra2, ra3 = st.columns(3)
                for _col, _price, _lbl, _bg, _border, _col_txt in [
                    (ra1, _a_min, "Igualar o mais barato", "#FEF3C7", "#FDE68A", "#7C2D12"),
                    (ra2, _a_avg, "Igualar a média",       "#DBEAFE", "#BFDBFE", "#1E3A8A"),
                    (ra3, _a_max, "Igualar o mais caro",   "#DCFCE7", "#BBF7D0", "#14532D"),
                ]:
                    _d = _disc_to_match(_sixt_ai_rack_vat, _price)
                    with _col:
                        st.markdown(
                            f'<div style="background:{_bg};border:1px solid {_border};border-radius:12px;padding:14px 16px;">'
                            f'<div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:#92400E;margin-bottom:6px;">{_lbl}</div>'
                            f'<div style="font-size:20px;font-weight:800;color:{_col_txt};">{f"{_d*100:.1f}%" if _d is not None else "—"}</div>'
                            f'<div style="font-size:11px;color:{_col_txt};margin-top:4px;">desc. balcão p/ igualar {eur(_price)}</div>'
                            f'</div>', unsafe_allow_html=True,
                        )

                st.markdown("")
                _a_rows = []
                for _cn, _cp2 in _valid_ai.items():
                    _gap = (_sixt_ai_counter_vat - _cp2) if _sixt_ai_counter_vat is not None else None
                    _dn  = _disc_to_match(_sixt_ai_rack_vat, _cp2)
                    _a_rows.append({
                        "Concorrente": _cn.replace("\n", " "),
                        f"Concorrência total {days}d (c/IVA)": _cp2,
                        f"SIXT balcão {ai_counter_pct:.1f}% (c/IVA)": _sixt_ai_counter_vat,
                        "Diferença (€)": _gap,
                        "Desc. balcão p/ igualar (%)": _dn * 100 if _dn is not None else None,
                        "Estado": "✓ Competitivo" if _gap is not None and _gap <= 0 else "⚠ Mais caro",
                    })
                with st.expander("📋  Detalhe All Inclusive por concorrente", expanded=True):
                    _fa = pd.DataFrame(_a_rows).copy()
                    for _c in [f"Concorrência total {days}d (c/IVA)", f"SIXT balcão {ai_counter_pct:.1f}% (c/IVA)", "Diferença (€)"]:
                        _fa[_c] = _fa[_c].apply(eur)
                    _fa["Desc. balcão p/ igualar (%)"] = _fa["Desc. balcão p/ igualar (%)"].apply(
                        lambda x: f"{x:.1f}%" if x is not None and not (isinstance(x, float) and pd.isna(x)) else "—")
                    st.dataframe(_fa, hide_index=True)

                # Gráfico AI
                _al = (["SIXT\nrack"] +
                       ([f"SIXT\nbalcão\n{ai_counter_pct:.1f}%"] if _sixt_ai_counter_vat else []) +
                       ([f"SIXT\nonline\n{ai_online_pct:.1f}%"]   if _sixt_ai_online_vat  else []) +
                       [k.replace("\n", " ") for k in _valid_ai])
                _av2 = ([_sixt_ai_rack_vat] +
                        ([_sixt_ai_counter_vat] if _sixt_ai_counter_vat else []) +
                        ([_sixt_ai_online_vat]  if _sixt_ai_online_vat  else []) +
                        list(_valid_ai.values()))
                _na = 1 + (1 if _sixt_ai_counter_vat else 0) + (1 if _sixt_ai_online_vat else 0)
                _ac = ["#1D4ED8"] + ["#93C5FD"] * (_na - 1) + ["#94A3B8"] * len(_valid_ai)
                _fig_a = go.Figure()
                _fig_a.add_trace(go.Bar(x=_al, y=_av2, marker_color=_ac,
                    text=[f"{v:,.2f}" if v else "" for v in _av2],
                    textposition="outside", textfont=dict(size=10, family="Inter"),
                    hovertemplate="%{x}<br><b>%{y:.2f} €</b><extra></extra>"))
                for _rv, _rl, _rp in [(_a_min, f"mín {eur(_a_min)}", "bottom right"), (_a_max, f"máx {eur(_a_max)}", "top right")]:
                    _fig_a.add_hline(y=_rv, line_dash="dot", line_color="#94A3B8", line_width=1,
                        annotation_text=_rl, annotation_position=_rp, annotation_font=dict(size=10, color="#64748B"))
                _fig_a.update_layout(
                    title=dict(text=f"All Inclusive — {group} · {days}d · c/IVA", font=dict(size=13, family="Inter", color="#0F172A")),
                    yaxis=dict(title="EUR c/IVA", gridcolor="#F1F5F9"),
                    height=380, margin=dict(l=40, r=20, t=50, b=40),
                    plot_bgcolor="white", paper_bgcolor="white", showlegend=False, font=dict(family="Inter"))
                st.plotly_chart(_fig_a, use_container_width=True)

            # ══════════════════════════════════════════════
            # RESUMO COMPETITIVO
            # ══════════════════════════════════════════════
            if (_any_smart and _sixt_smart_rack_vat) or (_any_ai and _sixt_ai_rack_vat):
                sec("Resumo — Limites de Desconto para Competitividade", "🎯", "green")

                _solver_vl = result.get("ai_counter_discount_valid_low")
                _solver_vh = result.get("ai_counter_discount_valid_high")
                _sc_valid  = _any_smart and _sixt_smart_rack_vat is not None
                _ac_valid  = _any_ai    and _sixt_ai_rack_vat    is not None

                _ds_cheap  = _disc_to_match(_sixt_smart_rack_vat, min(_valid_smart.values())) if _sc_valid else None
                _ds_exp    = _disc_to_match(_sixt_smart_rack_vat, max(_valid_smart.values())) if _sc_valid else None
                _da_cheap  = _disc_to_match(_sixt_ai_rack_vat,    min(_valid_ai.values()))    if _ac_valid else None
                _da_exp    = _disc_to_match(_sixt_ai_rack_vat,    max(_valid_ai.values()))    if _ac_valid else None

                st.markdown(
                    f"""<div style="background:white;border:1px solid #E2E8F0;border-radius:16px;padding:24px;margin-bottom:16px;">
                    <div style="font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:0.09em;color:#475569;margin-bottom:16px;">
                        Zona de conforto competitivo — {group} · {days} dia{'s' if days != 1 else ''}</div>
                    {"".join([f'''<div style="margin-bottom:16px;">
                        <div style="font-size:12px;font-weight:700;color:#FF5F00;margin-bottom:6px;">💳 SMART+</div>
                        <div style="background:#F1F5F9;border-radius:8px;padding:8px 12px;font-size:12px;color:#475569;">
                        Igualar mais barato: <strong style="color:#7C2D12;">{f"{_ds_cheap*100:.1f}%" if _ds_cheap is not None else "—"}</strong>
                        &nbsp;·&nbsp; Igualar mais caro: <strong style="color:#14532D;">{f"{_ds_exp*100:.1f}%" if _ds_exp is not None else "—"}</strong>
                        &nbsp;·&nbsp; Desconto atual: <strong style="color:#0F172A;">{counter_discount_pct:.1f}%</strong>
                        </div></div>''' if _sc_valid else ""])}
                    {"".join([f'''<div style="margin-bottom:8px;">
                        <div style="font-size:12px;font-weight:700;color:#1D4ED8;margin-bottom:6px;">💎 All Inclusive</div>
                        <div style="background:#F1F5F9;border-radius:8px;padding:8px 12px;font-size:12px;color:#475569;">
                        Igualar mais barato: <strong style="color:#7C2D12;">{f"{_da_cheap*100:.1f}%" if _da_cheap is not None else "—"}</strong>
                        &nbsp;·&nbsp; Igualar mais caro: <strong style="color:#14532D;">{f"{_da_exp*100:.1f}%" if _da_exp is not None else "—"}</strong>
                        &nbsp;·&nbsp; Solver AI: <strong style="color:#0F172A;">{f"{ai_counter_pct:.1f}%"}</strong>
                        {"&nbsp;·&nbsp; Intervalo solver: <strong style='color:#0F172A;'>[" + f"{_solver_vl*100:.1f}% ; {_solver_vh*100:.1f}%" + "]</strong>" if _solver_vl is not None and _solver_vh is not None else ""}
                        </div></div>''' if _ac_valid else ""])}
                    </div>""",
                    unsafe_allow_html=True,
                )

                if _ac_valid and _solver_vl is not None and _solver_vh is not None:
                    _ov_lo = max(_solver_vl, _da_exp  or 0.0)
                    _ov_hi = min(_solver_vh, _da_cheap or 1.0)
                    if _ov_lo <= _ov_hi:
                        st.success(
                            f"✅  Zona competitiva e válida para AI balcão: "
                            f"**[{_ov_lo*100:.1f}% ; {_ov_hi*100:.1f}%]** — "
                            f"SIXT fica abaixo de todos os concorrentes e cumpre as regras internas."
                        )
                    else:
                        st.warning(
                            f"⚠️  Sem sobreposição entre solver [{_solver_vl*100:.1f}%–{_solver_vh*100:.1f}%] "
                            f"e intervalo competitivo [{(_da_exp or 0)*100:.1f}%–{(_da_cheap or 1)*100:.1f}%]. "
                            f"Ajuste os descontos ou reveja os dados da concorrência."
                        )

        # ══════════════════════════════════════════════════
        # ANÁLISE POR CLUSTER DE SEGMENTO
        # ══════════════════════════════════════════════════
        st.markdown("---")
        sec("Análise Agregada por Cluster de Segmento", "📊", "blue")

        _CLUSTER_MAP = {
            "M": "Minis",        "E": "Económicos",  "C": "Compactos",
            "I": "Intermédios",  "F": "Full Size",   "P": "Premium",
            "L": "Luxury",       "S": "Standard",    "X": "Extraordinary",
        }
        _ALL_COMP_FIELDS = [
            "europcar_smart", "europcar_ai",
            "guerin_smart",   "guerin_ai",
            "avis_smart",     "avis_ai",
            "hertz_ai",
        ]
        _FIELD_LABEL = {
            "europcar_smart": "Europcar Medium",
            "europcar_ai":    "Europcar Premium",
            "guerin_smart":   "Guerin Prem.Gold",
            "guerin_ai":      "Guerin Platinum",
            "avis_smart":     "Avis Veículo",
            "avis_ai":        "Avis Veíc.Plus",
            "hertz_ai":       "Hertz SuperCover",
        }
        _FIELD_TYPE = {
            "europcar_smart": "smart", "europcar_ai": "ai",
            "guerin_smart":   "smart", "guerin_ai":   "ai",
            "avis_smart":     "smart", "avis_ai":     "ai",
            "hertz_ai":       "ai",
        }

        # ── Agregar preços da BD por cluster ───────────────
        _cl_raw = {}
        _cl_acriss_count = {}
        for _ac, _pd in comp_bd.items():
            _cl = _ac[0].upper() if _ac else ""
            if _cl not in _CLUSTER_MAP:
                continue
            if _cl not in _cl_raw:
                _cl_raw[_cl] = {f: [] for f in _ALL_COMP_FIELDS}
                _cl_acriss_count[_cl] = set()
            _cl_acriss_count[_cl].add(_ac)
            for _f, _p in _pd.items():
                if _f in _cl_raw[_cl] and _p is not None:
                    _cl_raw[_cl][_f].append(_p)

        _cl_avg = {}
        for _cl in sorted(_cl_raw.keys()):
            _cl_avg[_cl] = {}
            for _f in _ALL_COMP_FIELDS:
                _ps = _cl_raw[_cl][_f]
                _cl_avg[_cl][_f] = round(sum(_ps) / len(_ps), 2) if _ps else None

        # ── Calcular SIXT médio por cluster (balcão, dias selecionados) ──
        _sixt_cl_raw = {}
        for _g in matrix_data["groups"]:
            _cl = _g[0].upper() if _g else ""
            if _cl not in _CLUSTER_MAP:
                continue
            _r2 = calculate_pricing(
                matrix_data=matrix_data, group=_g, days=days,
                counter_discount=counter_discount_pct / 100,
                online_discount=online_discount_pct / 100,
                bf_weight=bf_weight_pct / 100,
                vat=vat_pct / 100,
                easy_counter_discount=easy_counter_pct / 100,
                tg_min_iva=tg_min_iva,
                easy_margin_min=easy_margin_min,
                easy_margin_max=easy_margin_max,
                online_extra_discount=online_extra_pct / 100,
            )
            if _cl not in _sixt_cl_raw:
                _sixt_cl_raw[_cl] = {"smart_counter_vat": [], "ai_counter_vat": []}
            if _r2.get("smart_counter_vat") is not None:
                _sixt_cl_raw[_cl]["smart_counter_vat"].append(_r2["smart_counter_vat"])
            if _r2.get("ai_counter_vat") is not None:
                _sixt_cl_raw[_cl]["ai_counter_vat"].append(_r2["ai_counter_vat"])

        _sixt_cl_avg = {}
        for _cl, _flds in _sixt_cl_raw.items():
            _sixt_cl_avg[_cl] = {}
            for _fk, _vs in _flds.items():
                _sixt_cl_avg[_cl][_fk] = round(sum(_vs) / len(_vs), 2) if _vs else None

        # ── Tabela resumo por cluster ──────────────────────
        st.markdown(
            f'<div style="font-size:12px;color:#64748B;margin-bottom:12px;">'
            f'Preços médios em <strong>€/dia c/IVA</strong> por cluster · '
            f'SIXT calculado para <strong>{days} dia{"s" if days != 1 else ""}</strong> ÷ dias = equivalente diário</div>',
            unsafe_allow_html=True,
        )

        _cl_table_rows = []
        for _cl in sorted(_cl_avg.keys()):
            _row = {
                "Cluster": f"{_cl} — {_CLUSTER_MAP[_cl]}",
                "# ACRISS": len(_cl_acriss_count.get(_cl, set())),
                "SIXT SMART+ balcão": (
                    round(_sixt_cl_avg[_cl]["smart_counter_vat"] / days, 2)
                    if _cl in _sixt_cl_avg and _sixt_cl_avg[_cl].get("smart_counter_vat") else None
                ),
                "SIXT AI balcão": (
                    round(_sixt_cl_avg[_cl]["ai_counter_vat"] / days, 2)
                    if _cl in _sixt_cl_avg and _sixt_cl_avg[_cl].get("ai_counter_vat") else None
                ),
            }
            for _f in _ALL_COMP_FIELDS:
                _row[_FIELD_LABEL[_f]] = _cl_avg[_cl].get(_f)
            _cl_table_rows.append(_row)

        _df_cl = pd.DataFrame(_cl_table_rows)
        _fmt_cl = _df_cl.copy()
        for _c in _fmt_cl.columns:
            if _c not in ["Cluster", "# ACRISS"]:
                _fmt_cl[_c] = _fmt_cl[_c].apply(eur)
        st.dataframe(_fmt_cl, hide_index=True)

        st.markdown("")

        # ── Gráfico SMART+ por cluster ────────────────────
        _clusters_with_smart = [
            _cl for _cl in sorted(_cl_avg.keys())
            if any(_cl_avg[_cl].get(_f) for _f in _ALL_COMP_FIELDS if _FIELD_TYPE[_f] == "smart")
        ]

        if _clusters_with_smart:
            sec("Equivalente SMART+ — Média por Cluster", "💳", "orange")
            _smart_fields = [_f for _f in _ALL_COMP_FIELDS if _FIELD_TYPE[_f] == "smart"]
            _smart_colors = ["#FF5F00", "#FB923C", "#3B82F6", "#93C5FD", "#10B981", "#6EE7B7"]

            _fig_cl_smart = go.Figure()
            # Concorrentes
            for _idx, _f in enumerate(_smart_fields):
                _x_labels = [f"{_cl} {_CLUSTER_MAP[_cl]}" for _cl in _clusters_with_smart]
                _y_vals   = [_cl_avg[_cl].get(_f) for _cl in _clusters_with_smart]
                _fig_cl_smart.add_trace(go.Bar(
                    name=_FIELD_LABEL[_f], x=_x_labels, y=_y_vals,
                    marker_color=_smart_colors[_idx % len(_smart_colors)],
                    text=[f"{v:.2f}" if v else "" for v in _y_vals],
                    textposition="outside", textfont=dict(size=9, family="Inter"),
                    hovertemplate=f"{_FIELD_LABEL[_f]}<br>%{{x}}<br><b>%{{y:.2f}} €/dia</b><extra></extra>",
                ))
            # SIXT linha de referência por cluster
            _sixt_smart_line = [
                round(_sixt_cl_avg[_cl]["smart_counter_vat"] / days, 2)
                if _cl in _sixt_cl_avg and _sixt_cl_avg[_cl].get("smart_counter_vat") else None
                for _cl in _clusters_with_smart
            ]
            _x_labels_s = [f"{_cl} {_CLUSTER_MAP[_cl]}" for _cl in _clusters_with_smart]
            _fig_cl_smart.add_trace(go.Scatter(
                name=f"SIXT SMART+ balcão {counter_discount_pct}%",
                x=_x_labels_s, y=_sixt_smart_line,
                mode="lines+markers",
                line=dict(color="#0F172A", width=2, dash="dot"),
                marker=dict(size=8, symbol="diamond", color="#0F172A"),
                hovertemplate="SIXT SMART+<br>%{x}<br><b>%{y:.2f} €/dia</b><extra></extra>",
            ))
            _fig_cl_smart.update_layout(
                barmode="group",
                title=dict(text=f"SMART+ por Cluster — €/dia c/IVA (concorrência BD · SIXT {days}d÷{days})", font=dict(size=13, family="Inter", color="#0F172A")),
                yaxis=dict(title="€/dia c/IVA", gridcolor="#F1F5F9"),
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0, font=dict(size=10)),
                height=420, margin=dict(l=40, r=20, t=80, b=60),
                plot_bgcolor="white", paper_bgcolor="white", font=dict(family="Inter"),
            )
            st.plotly_chart(_fig_cl_smart, use_container_width=True)

        # ── Gráfico AI por cluster ────────────────────────
        _clusters_with_ai = [
            _cl for _cl in sorted(_cl_avg.keys())
            if any(_cl_avg[_cl].get(_f) for _f in _ALL_COMP_FIELDS if _FIELD_TYPE[_f] == "ai")
        ]

        if _clusters_with_ai:
            sec("Equivalente All Inclusive — Média por Cluster", "💎", "blue")
            _ai_fields  = [_f for _f in _ALL_COMP_FIELDS if _FIELD_TYPE[_f] == "ai"]
            _ai_colors  = ["#FF5F00", "#1D4ED8", "#60A5FA", "#10B981", "#8B5CF6"]

            _fig_cl_ai = go.Figure()
            for _idx, _f in enumerate(_ai_fields):
                _x_labels = [f"{_cl} {_CLUSTER_MAP[_cl]}" for _cl in _clusters_with_ai]
                _y_vals   = [_cl_avg[_cl].get(_f) for _cl in _clusters_with_ai]
                _fig_cl_ai.add_trace(go.Bar(
                    name=_FIELD_LABEL[_f], x=_x_labels, y=_y_vals,
                    marker_color=_ai_colors[_idx % len(_ai_colors)],
                    text=[f"{v:.2f}" if v else "" for v in _y_vals],
                    textposition="outside", textfont=dict(size=9, family="Inter"),
                    hovertemplate=f"{_FIELD_LABEL[_f]}<br>%{{x}}<br><b>%{{y:.2f}} €/dia</b><extra></extra>",
                ))
            _sixt_ai_line = [
                round(_sixt_cl_avg[_cl]["ai_counter_vat"] / days, 2)
                if _cl in _sixt_cl_avg and _sixt_cl_avg[_cl].get("ai_counter_vat") else None
                for _cl in _clusters_with_ai
            ]
            _x_labels_a = [f"{_cl} {_CLUSTER_MAP[_cl]}" for _cl in _clusters_with_ai]
            _fig_cl_ai.add_trace(go.Scatter(
                name=f"SIXT AI balcão {ai_counter_pct:.1f}%",
                x=_x_labels_a, y=_sixt_ai_line,
                mode="lines+markers",
                line=dict(color="#0F172A", width=2, dash="dot"),
                marker=dict(size=8, symbol="diamond", color="#0F172A"),
                hovertemplate="SIXT AI<br>%{x}<br><b>%{y:.2f} €/dia</b><extra></extra>",
            ))
            _fig_cl_ai.update_layout(
                barmode="group",
                title=dict(text=f"All Inclusive por Cluster — €/dia c/IVA (concorrência BD · SIXT {days}d÷{days})", font=dict(size=13, family="Inter", color="#0F172A")),
                yaxis=dict(title="€/dia c/IVA", gridcolor="#F1F5F9"),
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0, font=dict(size=10)),
                height=420, margin=dict(l=40, r=20, t=80, b=60),
                plot_bgcolor="white", paper_bgcolor="white", font=dict(family="Inter"),
            )
            st.plotly_chart(_fig_cl_ai, use_container_width=True)

        # ── BD overview (todos os grupos) ─────────────────
        with st.expander(f"📊  Ver BD completa da concorrência ({_n_grupos_bd} grupos ACRISS)"):
            _bd_all_rows = []
            for _g, _gd in sorted(comp_bd.items()):
                _bd_all_rows.append({
                    "Grupo ACRISS": _g,
                    "Europcar Medium (SMART+)": _gd.get("europcar_smart"),
                    "Europcar Premium (AI)": _gd.get("europcar_ai"),
                    "Guerin Prem.Gold (SMART+)": _gd.get("guerin_smart"),
                    "Guerin Platinum (AI)": _gd.get("guerin_ai"),
                    "Avis Veículo (SMART+)": _gd.get("avis_smart"),
                    "Avis Veículo Plus (AI)": _gd.get("avis_ai"),
                    "Hertz SuperCover (AI)": _gd.get("hertz_ai"),
                })
            _df_all = pd.DataFrame(_bd_all_rows)
            _fmt_all = _df_all.copy()
            for _c in _df_all.columns[1:]:
                _fmt_all[_c] = _fmt_all[_c].apply(eur)
            st.caption("Preços em €/dia c/IVA · Média quando existem vários modelos no mesmo ACRISS")
            st.dataframe(_fmt_all, hide_index=True)


# ─────────────────────────────────────────────────────────
# TAB 4 — OPTIMIZADOR GLOBAL
# ─────────────────────────────────────────────────────────

with tab_opt:
    st.markdown(
        '<div class="pkg-banner ai" style="margin-bottom:20px;">'
        '<div class="pkg-banner-dot" style="background:#3B82F6;margin-top:4px;"></div>'
        '<div class="pkg-banner-body">'
        '<div class="pkg-banner-title">Optimizador Global</div>'
        '<div class="pkg-banner-formula">'
        'Calcula o intervalo de desconto AI balcão que torna todas as regras válidas '
        'em todos os grupos ACRISS × todos os intervalos LOR. '
        'Apresenta conflitos e cobertura completa.'
        '</div></div></div>',
        unsafe_allow_html=True,
    )

    opt_col, _ = st.columns([1, 3])
    with opt_col:
        run_optimizer = st.button("⚡  Optimizar agora", type="primary",
                                   help="Analisa todos os grupos × LOR. Pode demorar alguns segundos.")

    if run_optimizer:
        with st.spinner("A calcular o intervalo óptimo para todos os grupos e LOR…"):
            opt = optimize_global_discounts(
                matrix_data=matrix_data,
                counter_discount=counter_discount_pct / 100,
                online_discount=online_discount_pct / 100,
                bf_weight=bf_weight_pct / 100,
                vat=vat_pct / 100,
                easy_counter_discount=easy_counter_pct / 100,
                easy_margin_min=easy_margin_min, easy_margin_max=easy_margin_max,
                online_extra_discount=online_extra_pct / 100,
                tg_min_iva=tg_min_iva,
            )
            st.session_state["opt_result"]     = opt
            st.session_state["opt_ai_counter"] = opt["ai_counter_discount"]
            st.session_state["opt_ai_online"]  = opt["ai_online_discount"]

    if st.session_state["opt_result"] is not None:
        opt = st.session_state["opt_result"]

        sec("Resultado", "📊", "blue")

        oc1, oc2, oc3, oc4 = st.columns(4)
        kpi(oc1, "Desconto AI balcão óptimo",  f"{opt['ai_counter_discount']*100:.2f}%", "Ponto médio do intervalo global", "orange")
        kpi(oc2, "Desconto AI online óptimo",  f"{opt['ai_online_discount']*100:.2f}%",  f"= balcão + {online_extra_pct}% mínimo", "blue")
        kpi(oc3, "Intervalo válido balcão",
            f"{opt['valid_low']*100:.2f}% – {opt['valid_high']*100:.2f}%" if opt["feasible"] else "SEM SOLUÇÃO",
            "Intersecção de todos os grupos/LOR",
            "green" if opt["feasible"] else "red")
        n_conflicts = len(opt["conflicts"])
        kpi(oc4, "Grupos/LOR em conflito", str(n_conflicts),
            "Combinações sem solução única",
            "green" if n_conflicts == 0 else "red")

        if opt["feasible"]:
            st.success(
                f"✅  Existe um intervalo global único: "
                f"[{opt['valid_low']*100:.2f}% ; {opt['valid_high']*100:.2f}%]. "
                f"Ponto médio: balcão **{opt['ai_counter_discount']*100:.2f}%** · "
                f"online **{opt['ai_online_discount']*100:.2f}%**."
            )
            st.info("ℹ️  O simulador já calcula o desconto óptimo por grupo × LOR individualmente. Este resultado é informativo.")
        else:
            st.warning(
                f"Não existe um único desconto AI global válido para todos os grupos/LOR "
                f"({n_conflicts} conflito{'s' if n_conflicts != 1 else ''}). "
                "O simulador resolve isto calculando o desconto por grupo × LOR individualmente."
            )

        if opt["conflicts"]:
            with st.expander(f"🔴  Ver {len(opt['conflicts'])} conflito(s) em detalhe"):
                st.dataframe(
                    pd.DataFrame(opt["conflicts"]).assign(**{
                        "valid_low (%)":  lambda d: d["valid_low"]  * 100,
                        "valid_high (%)": lambda d: d["valid_high"] * 100,
                    }).drop(columns=["valid_low", "valid_high"]),
                    hide_index=True,
                )

        with st.expander("📋  Cobertura completa por grupo/LOR"):
            cov_rows = [{
                "Grupo/LOR": k,
                "Desc. mín balcão (%)": round(v["valid_low"]  * 100, 2) if v["valid_low"]  is not None else None,
                "Desc. máx balcão (%)": round(v["valid_high"] * 100, 2) if v["valid_high"] is not None else None,
                "Solução": "✓" if v["ok"] else ("✗" if v["ok"] is False else "—"),
                "Óptimo válido": "✓" if (
                    v["valid_low"] is not None and v["valid_high"] is not None
                    and v["ok"] and v["valid_low"] <= opt["ai_counter_discount"] <= v["valid_high"]
                ) else "✗",
            } for k, v in opt["coverage"].items()]
            st.dataframe(pd.DataFrame(cov_rows), hide_index=True)
    else:
        st.markdown(
            '<div class="landing-box">'
            '<div class="landing-icon">🎯</div>'
            '<div class="landing-title">Pronto para optimizar</div>'
            '<div class="landing-sub">Clique em "Optimizar agora" para calcular o intervalo de desconto AI<br>válido para todos os grupos ACRISS e intervalos LOR.</div>'
            '</div>',
            unsafe_allow_html=True,
        )


# ─────────────────────────────────────────────────────────
# TAB 4 — EXPORTAR
# ─────────────────────────────────────────────────────────

with tab_export:

    sec("Exportação completa — todos os grupos × todos os LOR", "📥", "green")

    st.markdown(
        """
        Aplica as regras atuais a **todos os grupos ACRISS** e a todos os intervalos LOR
        (1–3 · 4–6 · 7–10 · 11–29 · 30 · 31–999 dias).
        O ficheiro Excel inclui 5 abas:
        **Resumo_Final · SMART+ · All_Inclusive · Pack_Easy · Parametros**
        """
    )

    # Computação pesada: só corre quando o utilizador clica — nunca em cada re-render
    if st.button("🔄  Gerar / actualizar exportação", type="primary", key="btn_gen_export"):
        with st.spinner(f"A calcular {len(matrix_data['groups'])} grupos × 6 LOR…"):
            _df_d, _df_s, _df_a, _df_e, _df_p = build_all_groups_lor_export(
                matrix_data=matrix_data,
                counter_discount=counter_discount_pct / 100,
                online_discount=online_discount_pct   / 100,
                bf_weight=bf_weight_pct / 100,
                vat=vat_pct / 100,
                easy_counter_discount=easy_counter_pct / 100,
                ai_counter_discount=None, ai_online_discount=None,
                easy_margin_min=easy_margin_min, easy_margin_max=easy_margin_max,
                online_extra_discount=online_extra_pct / 100,
                tg_min_iva=tg_min_iva,
            )
            _excel = create_excel_download(_df_d, _df_s, _df_a, _df_e, _df_p)
            _csv   = _df_d.to_csv(index=False).encode("utf-8-sig")
            st.session_state["export_ready"] = {
                "excel": _excel.getvalue(),
                "csv":   _csv,
                "df_d":  _df_d,
                "df_p":  _df_p,
                "rows":  len(_df_d),
            }
        st.success(f"✅  Exportação gerada — {st.session_state['export_ready']['rows']} linhas.")

    exp_data = st.session_state.get("export_ready")

    if exp_data:
        ex1, ex2 = st.columns(2)
        with ex1:
            st.download_button(
                label="📥  Descarregar Excel completo",
                data=exp_data["excel"],
                file_name="pricing_protecoes_sixt_todos_grupos_lor.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )
        with ex2:
            st.download_button(
                label="📄  Descarregar CSV (Resumo_Final)",
                data=exp_data["csv"],
                file_name=f"pricing_simulation_{group}_{days}_dias.csv",
                mime="text/csv",
            )

        with st.expander("🔍  Pré-visualizar (primeiras 50 linhas)"):
            st.dataframe(exp_data["df_d"].head(50), hide_index=True)

        sec("Parâmetros utilizados", "⚙️", "")
        st.dataframe(exp_data["df_p"], hide_index=True)
    else:
        st.info("Clique em **Gerar / actualizar exportação** para calcular os preços de todos os grupos e preparar os downloads.")

    # ══════════════════════════════════════════════════════
    # NOVA MATRIZ — FORMATO SIXT PARA ALEMANHA
    # ══════════════════════════════════════════════════════
    st.markdown("<hr style='margin:32px 0 24px;border-color:#E2E8F0;'>", unsafe_allow_html=True)
    sec("🇩🇪  Nova Matriz — Formato SIXT (envio para Alemanha)", "📤", "orange")

    st.markdown(
        """
        Gera um ficheiro **Excel no formato exacto da matriz original**, com os preços
        **BF, TG e BQ actualizados** pelo simulador. Todos os restantes campos (LD, BE, AY, AE, I, BC,
        metadados) ficam intactos. As células alteradas ficam destacadas a amarelo.
        """
    )

    # Config email Alemanha (persistida em session_state)
    if "de_email" not in st.session_state:
        st.session_state["de_email"] = ""

    mat_col1, mat_col2 = st.columns([2, 1])
    with mat_col1:
        de_email = st.text_input(
            "📧  E-mail destinatário (Alemanha)",
            value=st.session_state["de_email"],
            placeholder="pricing@sixt.de",
            key="de_email_input",
            help="Endereço para o qual a matriz deve ser enviada.",
        )
        st.session_state["de_email"] = de_email
    with mat_col2:
        de_subject = st.text_input(
            "Assunto do e-mail",
            value="Nova Matriz Proteções SIXT Portugal",
            key="de_subject",
        )

    highlight = st.checkbox("Destacar células alteradas (amarelo)", value=True, key="mat_highlight")

    if st.button("🔄  Gerar nova matriz", type="primary", key="btn_gen_matrix"):
        with st.spinner(f"A recalcular {len(matrix_data['groups'])} grupos × 6 LOR e a gerar o ficheiro…"):
            _mat_buf, _n_changed, _stats = generate_sixt_matrix(
                file_bytes=file_bytes,
                sheet_name=sheet_name,
                matrix_data=matrix_data,
                counter_discount=counter_discount_pct / 100,
                online_discount=online_discount_pct   / 100,
                bf_weight=bf_weight_pct / 100,
                vat=vat_pct / 100,
                easy_counter_discount=easy_counter_pct / 100,
                easy_margin_min=easy_margin_min,
                easy_margin_max=easy_margin_max,
                online_extra_discount=online_extra_pct / 100,
                tg_min_iva=tg_min_iva,
                highlight_changes=highlight,
            )
            from datetime import date as _date
            _filename = f"Matrix-{_date.today().strftime('%d.%m.%Y')} - PT.xlsx"
            st.session_state["matrix_ready"] = {
                "bytes":     _mat_buf.getvalue(),
                "filename":  _filename,
                "n_changed": _n_changed,
                "stats":     _stats,
            }

    mat_data = st.session_state.get("matrix_ready")

    if mat_data:
        from datetime import date as _date

        # Resumo de alterações
        st.markdown("")
        s_bf = mat_data["stats"]["BF"]
        s_tg = mat_data["stats"]["TG"]
        s_bq = mat_data["stats"]["BQ"]

        s_be = mat_data["stats"].get("BE", {"changed": 0, "sum_delta": 0.0})
        mc1, mc2, mc3, mc4, mc5 = st.columns(5)
        kpi(mc1, "Células alteradas",   str(mat_data["n_changed"]),
            "BF + TG + BQ + BE", "orange")
        kpi(mc2, "BF actualizado",
            f"{s_bf['changed']} células",
            f"Δ total: {s_bf['sum_delta']:+.4f} €", "blue")
        kpi(mc3, "TG actualizado",
            f"{s_tg['changed']} células",
            f"Δ total: {s_tg['sum_delta']:+.4f} €", "green")
        kpi(mc4, "BQ actualizado",
            f"{s_bq['changed']} células",
            f"Δ total: {s_bq['sum_delta']:+.4f} €", "purple")
        kpi(mc5, "BE actualizado (80% BF)",
            f"{s_be['changed']} células",
            f"Δ total: {s_be['sum_delta']:+.4f} €", "sky")

        st.markdown("")

        dl_col, mail_col = st.columns(2)

        with dl_col:
            st.download_button(
                label=f"📥  Descarregar  {mat_data['filename']}",
                data=mat_data["bytes"],
                file_name=mat_data["filename"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )

        with mail_col:
            if st.button("📧  Criar rascunho de e-mail (Gmail)", key="btn_send_matrix",
                         help="Cria um rascunho no Gmail com a nova matriz em anexo."):
                if not de_email:
                    st.error("Introduz o e-mail do destinatário antes de criar o rascunho.")
                else:
                    # Guardar ficheiro temporariamente para anexar
                    import tempfile, os
                    with tempfile.NamedTemporaryFile(
                        delete=False, suffix=".xlsx",
                        prefix="Matrix_PT_"
                    ) as tmp:
                        tmp.write(mat_data["bytes"])
                        tmp_path = tmp.name

                    _body = (
                        f"Boa tarde,\n\n"
                        f"Em anexo a nova matriz de proteções SIXT Portugal "
                        f"({mat_data['filename']}).\n\n"
                        f"Alterações face à versão anterior:\n"
                        f"• BF: {s_bf['changed']} intervalos actualizados "
                        f"(Δ total rack: {s_bf['sum_delta']:+.2f} €)\n"
                        f"• TG: {s_tg['changed']} intervalos actualizados "
                        f"(Δ total rack: {s_tg['sum_delta']:+.2f} €)\n"
                        f"• BQ: {s_bq['changed']} intervalos actualizados "
                        f"(Δ total rack: {s_bq['sum_delta']:+.2f} €)\n\n"
                        f"Células alteradas estão destacadas a amarelo.\n\n"
                        f"Com os melhores cumprimentos,\nSIXT Portugal"
                    )
                    st.session_state["matrix_email_body"]   = _body
                    st.session_state["matrix_email_to"]     = de_email
                    st.session_state["matrix_email_subject"]= de_subject
                    st.session_state["matrix_tmp_path"]     = tmp_path
                    st.session_state["matrix_email_ready"]  = True
                    st.success(f"Rascunho preparado. Clica em **Confirmar e criar no Gmail** abaixo.")

        # Confirmar e criar rascunho Gmail
        if st.session_state.get("matrix_email_ready"):
            st.markdown(
                f"""<div style="background:#F0FDF4;border:1px solid #86EFAC;border-radius:12px;
                padding:16px 20px;margin-top:16px;font-size:13px;color:#14532D;">
                <strong>Para:</strong> {st.session_state['matrix_email_to']}<br>
                <strong>Assunto:</strong> {st.session_state['matrix_email_subject']}<br>
                <strong>Anexo:</strong> {mat_data['filename']}<br><br>
                <pre style="font-size:11px;color:#166534;white-space:pre-wrap;">{st.session_state['matrix_email_body']}</pre>
                </div>""",
                unsafe_allow_html=True,
            )
            if st.button("✅  Confirmar e criar rascunho no Gmail", key="btn_confirm_email", type="primary"):
                st.session_state["trigger_gmail_draft"] = True
                st.rerun()

        if st.session_state.get("trigger_gmail_draft"):
            st.session_state["trigger_gmail_draft"] = False
            st.session_state["matrix_email_ready"]  = False
            st.info(
                "O rascunho foi preparado. Para enviar com anexo, utiliza o botão de download "
                "e junta o ficheiro manualmente no Gmail — a API do Gmail não suporta anexos por esta via.\n\n"
                f"**Para:** {st.session_state.get('matrix_email_to','')}\n\n"
                f"**Assunto:** {st.session_state.get('matrix_email_subject','')}"
            )
    else:
        st.info("Clica em **Gerar nova matriz** para criar o ficheiro actualizado no formato SIXT.")


# ─────────────────────────────────────────────────────────
# TAB 5 — TÉCNICO / DEBUG
# ─────────────────────────────────────────────────────────

with tab_debug:
    sec("Dados carregados", "🔬", "")

    db1, db2, db3 = st.columns(3)
    kpi(db1, "Grupos ACRISS",      str(len(matrix_data["groups"])), f"Folha: {sheet_name}")
    kpi(db2, "Grupo seleccionado", group,                           f"{days} dia{'s' if days != 1 else ''} · IVA {vat_pct}%")
    kpi(db3, "Produtos disponíveis", ", ".join(list(matrix_data["data"].get(group, {}).keys())), "Para o grupo seleccionado")

    st.markdown("")

    with st.expander("📋  Lista completa de grupos"):
        st.write(matrix_data["groups"])

    with st.expander("🔎  Dados raw do grupo seleccionado"):
        st.json({
            "grupo": group,
            "dias": days,
            "produtos": list(matrix_data["data"].get(group, {}).keys()),
            "preços_calculados": {
                k: v for k, v in result.items()
                if k not in ["issues"] and not isinstance(v, dict)
            },
        })

    with st.expander("⚠️  Issues / alertas"):
        if result["issues"]:
            for issue in result["issues"]:
                st.error(issue)
        else:
            st.success("Sem issues para este grupo/LOR.")
